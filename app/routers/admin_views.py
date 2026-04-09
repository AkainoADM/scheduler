import os
from datetime import date, datetime
from typing import List
from fastapi import APIRouter, Depends, Form, Request, BackgroundTasks, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import httpx

from app.core.database import get_db
from app.services import (
    group as group_service,
    teacher as teacher_service,
    audience as audience_service,
    faculty as faculty_service,
    building as building_service,
    calendar as calendar_service,
    user_type as user_type_service,
    subdivision as subdivision_service,
    role as role_service,
    permission as permission_service,
    time_slot as time_slot_service,
    subject as subject_service
)
from app.services import calendar_sync
from app.schemas.group import GroupCreate, GroupUpdate
from app.schemas.teacher import TeacherCreate, TeacherUpdate
from app.schemas.audience import AudienceCreate, AudienceUpdate
from app.schemas.building import BuildingCreate, BuildingUpdate
from app.schemas.calendar import CalendarCreate, CalendarUpdate
from app.schemas.faculty import FacultyCreate, FacultyUpdate
from app.schemas.user_type import UserTypeCreate, UserTypeUpdate
from app.schemas.subdivision import SubdivisionCreate, SubdivisionUpdate
from app.schemas.role import RoleCreate, RoleUpdate
from app.schemas.permission import PermissionCreate, PermissionUpdate
from app.schemas.time_slot import TimeSlotCreate, TimeSlotUpdate
from app.schemas.subject import SubjectCreate, SubjectUpdate

from app.services import teacher_group as teacher_group_service
from app.services import teacher_subject as teacher_subject_service
from app.services import audience_subject as audience_subject_service
from app.services import group_subject as group_subject_service
from app.schemas.teacher_group import TeacherGroupCreate
from app.schemas.teacher_subject import TeacherSubjectCreate, TeacherSubjectUpdate
from app.schemas.audience_subject import AudienceSubjectCreate
from app.schemas.group_subject import GroupSubjectCreate

from app.services import user as user_service
from app.schemas.user import UserCreate, UserUpdate

from sqlalchemy.orm import selectinload
from app.models.reference import ScheduleItem, FinalScheduleItem, Lesson, Subject, TimeSlot, Audience

router = APIRouter(prefix="/admin", tags=["Admin Pages"])

# ---------- Вспомогательные функции для проверки зависимостей ----------
async def has_dependent_groups(db: AsyncSession, faculty_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import Group
    result = await db.execute(select(func.count()).select_from(Group).where(Group.faculty_id == faculty_id))
    return result.scalar() > 0

async def has_dependent_audiences(db: AsyncSession, building_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import Audience
    result = await db.execute(select(func.count()).select_from(Audience).where(Audience.building_id == building_id))
    return result.scalar() > 0

async def has_dependent_audience_subjects(db: AsyncSession, audience_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import op_audience_subject
    result = await db.execute(select(func.count()).select_from(op_audience_subject).where(op_audience_subject.c.audience_id == audience_id))
    return result.scalar() > 0

async def has_dependent_teacher_groups(db: AsyncSession, teacher_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import op_teachers_of_groups
    result = await db.execute(select(func.count()).select_from(op_teachers_of_groups).where(op_teachers_of_groups.c.teacher_id == teacher_id))
    return result.scalar() > 0

async def has_dependent_teacher_subjects(db: AsyncSession, teacher_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import op_teachers_of_pairs
    result = await db.execute(select(func.count()).select_from(op_teachers_of_pairs).where(op_teachers_of_pairs.c.teacher_id == teacher_id))
    return result.scalar() > 0

async def has_dependent_group_teachers(db: AsyncSession, group_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import op_teachers_of_groups
    result = await db.execute(select(func.count()).select_from(op_teachers_of_groups).where(op_teachers_of_groups.c.group_id == group_id))
    return result.scalar() > 0

async def has_dependent_group_subjects(db: AsyncSession, group_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import op_groups_of_pairs
    result = await db.execute(select(func.count()).select_from(op_groups_of_pairs).where(op_groups_of_pairs.c.group_id == group_id))
    return result.scalar() > 0

async def has_dependent_subject_teachers(db: AsyncSession, subject_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import op_teachers_of_pairs
    result = await db.execute(select(func.count()).select_from(op_teachers_of_pairs).where(op_teachers_of_pairs.c.subject_id == subject_id))
    return result.scalar() > 0

async def has_dependent_subject_audiences(db: AsyncSession, subject_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import op_audience_subject
    result = await db.execute(select(func.count()).select_from(op_audience_subject).where(op_audience_subject.c.subject_id == subject_id))
    return result.scalar() > 0

async def has_dependent_subject_groups(db: AsyncSession, subject_id: int) -> bool:
    from sqlalchemy import func
    from app.models.reference import op_groups_of_pairs
    result = await db.execute(select(func.count()).select_from(op_groups_of_pairs).where(op_groups_of_pairs.c.subject_id == subject_id))
    return result.scalar() > 0

# ---------- Общая функция экспорта в Excel ----------
def export_to_excel(data: list, columns: list, sheet_name: str, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(columns, 1):
            value = row_data.get(key)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ---------- Вспомогательная функция для импорта через API ----------
async def proxy_upload(entity: str, file: UploadFile, db: AsyncSession):
    api_url = f"http://localhost:8000/api/v1/{entity}/upload"
    contents = await file.read()
    files = {"file": (file.filename, contents, file.content_type)}
    async with httpx.AsyncClient() as client:
        response = await client.post(api_url, files=files)
        response.raise_for_status()
        return response.json()

# Добавьте в app/routers/admin_views.py (после импортов, но до router)
from app.models.reference import ScheduleItem, Lesson, Subject, TimeSlot, Audience, FinalScheduleItem

@router.get("/schedule-editor", response_class=HTMLResponse)
async def schedule_editor_page(db: AsyncSession = Depends(get_db)):
    # Загружаем все черновики с необходимыми связями
    items = await db.execute(
        select(ScheduleItem)
        .options(
            selectinload(ScheduleItem.lesson).selectinload(Lesson.subject),
            selectinload(ScheduleItem.time_slot),
            selectinload(ScheduleItem.audience)
        )
        .order_by(ScheduleItem.date, ScheduleItem.time_slot_id)
    )
    items = items.scalars().all()

    # Для выпадающих списков аудиторий и временных слотов
    audiences = await audience_service.get_all_audiences(db)
    time_slots = await time_slot_service.get_all_time_slots(db)

    # Формируем данные для таблицы
    rows = []
    for item in items:
        lesson = item.lesson
        subject = lesson.subject if lesson else None
        rows.append({
            "id": item.id,
            "date": item.date,
            "subject": subject.name if subject else "—",
            "time_slot": item.time_slot,
            "audience": item.audience,
            "lesson_id": lesson.id if lesson else None,
        })

    # Генерируем HTML
    html = f"""
    <html>
    <head>
        <title>Редактор черновика расписания</title>
        {_common_styles()}
        <style>
            .edit-form {{ display: inline; margin-left: 10px; }}
            select, input {{ padding: 4px; margin: 2px; }}
            button {{ margin: 2px; }}
        </style>
    </head>
    <body>
        <h1>Черновик расписания</h1>
        <div><a href="/admin/export/schedule_items" class="btn-excel">📎 Экспорт черновика в Excel</a></div>
        <div style="margin-top: 20px;">
            <form method="post" action="/admin/schedule/approve" onsubmit="return confirm('Утвердить расписание? Текущее опубликованное будет заменено.');">
                <button type="submit" class="btn" style="background: #28a745;">✅ Утвердить расписание</button>
            </form>
        </div>
        <table border="1" style="width:100%; margin-top:20px;">
            <thead>
                <tr><th>ID</th><th>Дата</th><th>Предмет</th><th>Время</th><th>Аудитория</th><th>Действия</th></tr>
            </thead>
            <tbody>
    """
    for row in rows:
        # Строим выпадающие списки для аудиторий и времени
        audience_options = ""
        for a in audiences:
            selected = "selected" if row["audience"] and a.id == row["audience"].id else ""
            audience_options += f'<option value="{a.id}" {selected}>{a.name}</option>'
        time_slot_options = ""
        for ts in time_slots:
            selected = "selected" if row["time_slot"] and ts.id == row["time_slot"].id else ""
            time_slot_options += f'<option value="{ts.id}" {selected}>{ts.name} ({ts.start_time}-{ts.end_time})</option>'

        html += f"""
            <tr>
                <td>{row["id"]}</td>
                <td><input type="date" name="date_{row["id"]}" value="{row["date"]}" form="edit_{row["id"]}" style="width:120px;"></td>
                <td>{row["subject"]}</td>
                <td><select name="time_slot_id_{row["id"]}" form="edit_{row["id"]}">{time_slot_options}</select></td>
                <td><select name="audience_id_{row["id"]}" form="edit_{row["id"]}">{audience_options}</select></td>
                <td>
                    <form id="edit_{row["id"]}" method="post" action="/admin/schedule/update/{row["id"]}" style="display:inline;">
                        <button type="submit">Сохранить</button>
                    </form>
                    <a href="/admin/schedule/delete/{row["id"]}" onclick="return confirm('Удалить занятие из черновика?');">Удалить</a>
                </td>
            </tr>
        """
    html += """
            </tbody>
        </table>
        <br><a href="/admin/">На главную админки</a>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

@router.post("/schedule/update/{item_id}", response_model=None)
async def update_schedule_item(
    item_id: int,
    date: date = Form(...),
    time_slot_id: int = Form(...),
    audience_id: int = Form(...),
    db: AsyncSession = Depends(get_db)
):
    item = await db.get(ScheduleItem, item_id)
    if item:
        item.date = date
        item.time_slot_id = time_slot_id
        item.audience_id = audience_id
        await db.commit()
    return RedirectResponse(url="/admin/schedule-editor", status_code=303)

@router.get("/schedule/delete/{item_id}", response_class=HTMLResponse)
async def confirm_delete_schedule_item(item_id: int, db: AsyncSession = Depends(get_db)):
    item = await db.get(ScheduleItem, item_id)
    if not item:
        return HTMLResponse("Запись не найдена", status_code=404)
    html = f"""
    <h1>Удалить занятие из черновика?</h1>
    <p>ID: {item.id}, Дата: {item.date}</p>
    <form method="post" action="/admin/schedule/delete/{item_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/schedule-editor">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/schedule/delete/{item_id}", response_model=None)
async def delete_schedule_item(item_id: int, db: AsyncSession = Depends(get_db)):
    await db.execute(delete(ScheduleItem).where(ScheduleItem.id == item_id))
    await db.commit()
    return RedirectResponse(url="/admin/schedule-editor", status_code=303)

@router.post("/schedule/approve", response_model=None)
async def approve_schedule(db: AsyncSession = Depends(get_db)):
    # Очищаем старую публикацию
    await db.execute(delete(FinalScheduleItem))
    # Копируем все черновики
    drafts = await db.execute(select(ScheduleItem))
    drafts = drafts.scalars().all()
    for d in drafts:
        final = FinalScheduleItem(
            lesson_id=d.lesson_id,
            time_slot_id=d.time_slot_id,
            audience_id=d.audience_id,
            date=d.date
        )
        db.add(final)
    await db.commit()
    return RedirectResponse(url="/admin/schedule-editor", status_code=303)

@router.get("/export/schedule_items")
async def export_schedule_items(db: AsyncSession = Depends(get_db)):
    items = await db.execute(
        select(ScheduleItem)
        .options(
            selectinload(ScheduleItem.lesson).selectinload(Lesson.subject),
            selectinload(ScheduleItem.time_slot),
            selectinload(ScheduleItem.audience)
        )
    )
    items = items.scalars().all()
    data = []
    for item in items:
        lesson = item.lesson
        subject = lesson.subject if lesson else None
        data.append({
            "ID": item.id,
            "Дата": item.date,
            "Предмет": subject.name if subject else "—",
            "Время": item.time_slot.name if item.time_slot else "—",
            "Аудитория": item.audience.name if item.audience else "—",
            "Закреплено": "Да" if item.is_pinned else "Нет"
        })
    columns = ["ID", "Дата", "Предмет", "Время", "Аудитория", "Закреплено"]
    return export_to_excel(data, columns, "Черновик расписания", "schedule_items.xlsx")

@router.get("/final-schedule", response_class=HTMLResponse)
async def final_schedule_page(db: AsyncSession = Depends(get_db)):
    items = await db.execute(
        select(FinalScheduleItem)
        .options(
            selectinload(FinalScheduleItem.lesson).selectinload(Lesson.subject),
            selectinload(FinalScheduleItem.time_slot),
            selectinload(FinalScheduleItem.audience)
        )
        .order_by(FinalScheduleItem.date, FinalScheduleItem.time_slot_id)
    )
    items = items.scalars().all()
    rows = []
    for item in items:
        lesson = item.lesson
        subject = lesson.subject if lesson else None
        rows.append({
            "date": item.date,
            "subject": subject.name if subject else "—",
            "time_slot": item.time_slot.name if item.time_slot else "—",
            "audience": item.audience.name if item.audience else "—",
        })
    html = f"""
    <html>
    <head><title>Опубликованное расписание</title>{_common_styles()}</head>
    <body>
        <h1>Опубликованное расписание</h1>
        <div><a href="/admin/export/final_schedule" class="btn-excel">📎 Экспорт в Excel</a></div>
        <table border="1" style="width:100%; margin-top:20px;">
            <thead><tr><th>Дата</th><th>Предмет</th><th>Время</th><th>Аудитория</th></tr></thead>
            <tbody>
    """
    for row in rows:
        html += f"<tr><td>{row['date']}</td><td>{row['subject']}</td><td>{row['time_slot']}</td><td>{row['audience']}</td></tr>"
    html += """
            </tbody>
        </table>
        <br><a href="/admin/">На главную админки</a>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

@router.get("/export/final_schedule")
async def export_final_schedule(db: AsyncSession = Depends(get_db)):
    items = await db.execute(
        select(FinalScheduleItem)
        .options(
            selectinload(FinalScheduleItem.lesson).selectinload(Lesson.subject),
            selectinload(FinalScheduleItem.time_slot),
            selectinload(FinalScheduleItem.audience)
        )
    )
    items = items.scalars().all()
    data = []
    for item in items:
        lesson = item.lesson
        subject = lesson.subject if lesson else None
        data.append({
            "Дата": item.date,
            "Предмет": subject.name if subject else "—",
            "Время": item.time_slot.name if item.time_slot else "—",
            "Аудитория": item.audience.name if item.audience else "—",
        })
    columns = ["Дата", "Предмет", "Время", "Аудитория"]
    return export_to_excel(data, columns, "Опубликованное расписание", "final_schedule.xlsx")

# ---------- Эндпоинт импорта ----------
@router.post("/import/{entity}")
async def import_entity(
    entity: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    try:
        result = await proxy_upload(entity, file, db)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ---------- Вспомогательные функции для рендеринга (с кнопками экспорта и импорта) ----------
def _common_styles():
    return """
    <style>
        .btn-excel {
            display: inline-block;
            background: #28a745;
            color: white;
            padding: 6px 12px;
            text-decoration: none;
            border-radius: 4px;
            margin-bottom: 15px;
        }
        .btn-import {
            background: #007bff;
            margin-left: 10px;
        }
        table { border-collapse: collapse; width: 100%; margin-top: 10px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        .import-form { display: none; margin: 10px 0; padding: 10px; border: 1px solid #ccc; background: #f9f9f9; }
    </style>
    """

def _import_script(entity: str):
    return f"""
    <script>
        function toggleImportForm_{entity}() {{
            var form = document.getElementById('importForm_{entity}');
            if (form.style.display === 'none' || form.style.display === '') {{
                form.style.display = 'block';
            }} else {{
                form.style.display = 'none';
            }}
        }}
        document.getElementById('uploadForm_{entity}').onsubmit = async function(e) {{
            e.preventDefault();
            const formData = new FormData(this);
            const response = await fetch('/admin/import/{entity}', {{ method: 'POST', body: formData }});
            const result = await response.json();
            alert(`Добавлено: ${{result.added}}, Обновлено: ${{result.updated}}\\nОшибки: ${{result.errors.length}}`);
            window.location.reload();
        }};
    </script>
    """

def render_groups_html(groups, faculties):
    html = f"<h1>Группы</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/groups' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_groups()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_groups' class='import-form'>"
    html += f"<form id='uploadForm_groups' action='/admin/import/groups' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_groups()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("groups")
    # Форма добавления
    html += "<form method='post' action='/admin/groups/add'>"
    html += "<input type='text' name='name' placeholder='Название группы' required>"
    html += "<select name='faculty_id' required><option value=''>Выберите факультет</option>"
    for fac in faculties:
        html += f"<option value='{fac.id}'>{fac.name}</option>"
    html += "</select>"
    html += "<input type='number' name='student_count' placeholder='Количество студентов'>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<form method='post' action='/admin/groups/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllGroups'> Выделить всё</label>"
    html += "<ul>"
    for g in groups:
        html += f"<li><input type='checkbox' name='ids' value='{g.id}'> "
        html += f"{g.name} (Факультет ID: {g.faculty_id}, Студентов: {g.student_count}) "
        html += f"<a href='/admin/groups/edit/{g.id}'>Редактировать</a> "
        html += f"<a href='/admin/groups/delete/{g.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllGroups');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_teachers_html(teachers):
    html = f"<h1>Преподаватели</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/teachers' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_teachers()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_teachers' class='import-form'>"
    html += f"<form id='uploadForm_teachers' action='/admin/import/teachers' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_teachers()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("teachers")
    html += "<form method='post' action='/admin/teachers/add'>"
    html += "<input type='text' name='login' placeholder='Логин' required>"
    html += "<input type='text' name='name' placeholder='ФИО' required>"
    html += "<input type='text' name='url' placeholder='Ссылка на страницу'>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<form method='post' action='/admin/teachers/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllTeachers'> Выделить всё</label>"
    html += "<ul>"
    for t in teachers:
        html += f"<li><input type='checkbox' name='ids' value='{t.id}'> "
        html += f"{t.name} (Логин: {t.login})"
        if t.url:
            html += f" – <a href='{t.url}'>страница</a>"
        html += f" <a href='/admin/teachers/edit/{t.id}'>Редактировать</a> "
        html += f"<a href='/admin/teachers/delete/{t.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllTeachers');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_audiences_html(audiences, buildings):
    html = f"<h1>Аудитории</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/audiences' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_audiences()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_audiences' class='import-form'>"
    html += f"<form id='uploadForm_audiences' action='/admin/import/audiences' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_audiences()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("audiences")
    html += "<form method='post' action='/admin/audiences/add'>"
    html += "<input type='text' name='name' placeholder='Номер/название' required>"
    html += "<select name='building_id'><option value=''>Не выбрано</option>"
    for b in buildings:
        html += f"<option value='{b.id}'>{b.name}</option>"
    html += "</select>"
    html += "<input type='number' name='capacity' placeholder='Вместимость'>"
    html += "<select name='type'><option value=''>Тип</option><option value='лекционная'>Лекционная</option><option value='практическая'>Практическая</option><option value='лабораторная'>Лабораторная</option><option value='компьютерный класс'>Компьютерный класс</option></select>"
    html += "<label><input type='checkbox' name='is_active' value='true' checked> Активна</label>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<form method='post' action='/admin/audiences/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllAudiences'> Выделить всё</label>"
    html += "<ul>"
    for a in audiences:
        html += f"<li><input type='checkbox' name='ids' value='{a.id}'> "
        html += f"{a.name} (Корпус ID: {a.building_id or 'не указан'}, Мест: {a.capacity or 'не указано'}, Тип: {a.type or 'не указан'}, Активна: {a.is_active}) "
        html += f"<a href='/admin/audiences/edit/{a.id}'>Редактировать</a> "
        html += f"<a href='/admin/audiences/delete/{a.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllAudiences');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_buildings_html(buildings):
    html = f"<h1>Здания</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/buildings' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_buildings()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_buildings' class='import-form'>"
    html += f"<form id='uploadForm_buildings' action='/admin/import/buildings' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_buildings()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("buildings")
    html += "<form method='post' action='/admin/buildings/add'>"
    html += "<input type='text' name='name' placeholder='Название здания' required>"
    html += "<input type='text' name='address' placeholder='Адрес'>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<form method='post' action='/admin/buildings/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllBuildings'> Выделить всё</label>"
    html += "<ul>"
    for b in buildings:
        html += f"<li><input type='checkbox' name='ids' value='{b.id}'> "
        html += f"{b.name} (Адрес: {b.address or 'не указан'}) "
        html += f"<a href='/admin/buildings/edit/{b.id}'>Редактировать</a> "
        html += f"<a href='/admin/buildings/delete/{b.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllBuildings');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_calendar_html(entries):
    html = f"<h1>Календарь (учебные дни, праздники)</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/calendar' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_calendar()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_calendar' class='import-form'>"
    html += f"<form id='uploadForm_calendar' action='/admin/import/calendar' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_calendar()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("calendar")
    html += "<h2>Добавить запись</h2>"
    html += "<form method='post' action='/admin/calendar/add'>"
    html += "<input type='date' name='date' required>"
    html += "<select name='is_working_day'><option value='true'>Рабочий день</option><option value='false'>Выходной/праздник</option></select>"
    html += "<input type='text' name='description' placeholder='Описание'>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<h2>Синхронизация с официальным календарем</h2>"
    html += "<form method='post' action='/admin/calendar/sync'>"
    html += "<input type='number' name='year' placeholder='Год (например, 2025)' required>"
    html += "<button type='submit'>Загрузить данные за год</button>"
    html += "</form>"
    html += "<h2>Список записей</h2>"
    html += "<form method='post' action='/admin/calendar/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllCalendar'> Выделить всё</label>"
    html += "<ul>"
    for e in entries:
        html += f"<li><input type='checkbox' name='ids' value='{e.id}'> "
        html += f"{e.date} – {'Рабочий' if e.is_working_day else 'Выходной'} "
        if e.description:
            html += f"— {e.description} "
        html += f"<a href='/admin/calendar/edit/{e.id}'>Редактировать</a> "
        html += f"<a href='/admin/calendar/delete/{e.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllCalendar');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_faculties_html(faculties):
    html = f"<h1>Факультеты</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/faculties' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_faculties()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_faculties' class='import-form'>"
    html += f"<form id='uploadForm_faculties' action='/admin/import/faculties' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_faculties()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("faculties")
    html += "<form method='post' action='/admin/faculties/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllFaculties'> Выделить всё</label>"
    html += "</form>"
    html += "<form method='post' action='/admin/faculties/add' style='margin-top: 20px;'>"
    html += "<input type='text' name='name' placeholder='Название' required>"
    html += "<input type='text' name='display_name' placeholder='Отображаемое имя'>"
    html += "<input type='text' name='short_display_name' placeholder='Короткое имя'>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<ul>"
    for f in faculties:
        html += f"<li><input type='checkbox' name='ids' value='{f.id}'> "
        html += f"{f.name} (ID: {f.id})"
        if f.display_name:
            html += f" – {f.display_name}"
        if f.short_display_name:
            html += f" ({f.short_display_name})"
        html += f" <a href='/admin/faculties/edit/{f.id}'>Редактировать</a> "
        html += f"<a href='/admin/faculties/delete/{f.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllFaculties');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_users_html(users, user_types):
    html = f"<h1>Пользователи</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/users' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_users()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_users' class='import-form'>"
    html += f"<form id='uploadForm_users' action='/admin/import/users' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_users()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("users")
    html += "<form method='post' action='/admin/users/add'>"
    html += "<input type='text' name='username' placeholder='Логин' required>"
    html += "<input type='email' name='email' placeholder='Email' required>"
    html += "<input type='password' name='password' placeholder='Пароль' required>"
    html += "<input type='text' name='full_name' placeholder='ФИО'>"
    html += "<select name='user_type_id'><option value=''>Тип пользователя</option>"
    for ut in user_types:
        html += f"<option value='{ut.id}'>{ut.name}</option>"
    html += "</select>"
    html += "<label><input type='checkbox' name='is_active' value='true' checked> Активен</label>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<form method='post' action='/admin/users/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllUsers'> Выделить всё</label>"
    html += "<ul>"
    for u in users:
        html += f"<li><input type='checkbox' name='ids' value='{u.id}'> "
        html += f"{u.username} ({u.email}) – {u.full_name or '—'}"
        if u.user_type:
            html += f" [{u.user_type.name}]"
        html += f" <a href='/admin/users/edit/{u.id}'>Редактировать</a> "
        html += f"<a href='/admin/users/delete/{u.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllUsers');
        if(selectAll) selectAll.addEventListener('change', function() {
            document.querySelectorAll('input[name="ids"]').forEach(cb => cb.checked = selectAll.checked);
        });
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Удалить выбранных пользователей?');
        }
    </script>
    """
    return html

def render_user_types_html(user_types):
    html = f"<h1>Типы пользователей</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/user_types' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_user_types()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_user_types' class='import-form'>"
    html += f"<form id='uploadForm_user_types' action='/admin/import/user_types' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_user_types()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("user_types")
    html += "<form method='post' action='/admin/user-types/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllUserTypes'> Выделить всё</label>"
    html += "<ul>"
    for ut in user_types:
        html += f"<li><input type='checkbox' name='ids' value='{ut.id}'> "
        html += f"{ut.name} (код: {ut.code})"
        if ut.description:
            html += f" – {ut.description}"
        html += f" <a href='/admin/user-types/edit/{ut.id}'>Редактировать</a> "
        html += f"<a href='/admin/user-types/delete/{ut.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<form method='post' action='/admin/user-types/add' style='margin-top: 20px;'>"
    html += "<input type='text' name='code' placeholder='Код (например, teacher)' required>"
    html += "<input type='text' name='name' placeholder='Название' required>"
    html += "<input type='text' name='description' placeholder='Описание'>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllUserTypes');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_subdivisions_html(subdivisions):
    html = f"<h1>Подразделения</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/subdivisions' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_subdivisions()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_subdivisions' class='import-form'>"
    html += f"<form id='uploadForm_subdivisions' action='/admin/import/subdivisions' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_subdivisions()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("subdivisions")
    html += "<form method='post' action='/admin/subdivisions/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllSubdivisions'> Выделить всё</label>"
    html += "<ul>"
    for s in subdivisions:
        html += f"<li><input type='checkbox' name='ids' value='{s.id}'> "
        html += f"{s.name}"
        html += f" <a href='/admin/subdivisions/edit/{s.id}'>Редактировать</a> "
        html += f"<a href='/admin/subdivisions/delete/{s.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<form method='post' action='/admin/subdivisions/add' style='margin-top: 20px;'>"
    html += "<input type='text' name='name' placeholder='Название' required>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllSubdivisions');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_roles_html(roles):
    html = f"<h1>Роли</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/roles' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_roles()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_roles' class='import-form'>"
    html += f"<form id='uploadForm_roles' action='/admin/import/roles' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_roles()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("roles")
    html += "<form method='post' action='/admin/roles/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllRoles'> Выделить всё</label>"
    html += "<ul>"
    for r in roles:
        html += f"<li><input type='checkbox' name='ids' value='{r.id}'> "
        html += f"{r.role_name}"
        if r.description:
            html += f" – {r.description}"
        html += f" <a href='/admin/roles/edit/{r.id}'>Редактировать</a> "
        html += f"<a href='/admin/roles/delete/{r.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<form method='post' action='/admin/roles/add' style='margin-top: 20px;'>"
    html += "<input type='text' name='role_name' placeholder='Название роли' required>"
    html += "<input type='text' name='description' placeholder='Описание'>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllRoles');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_permissions_html(permissions):
    html = f"<h1>Права</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/permissions' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_permissions()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_permissions' class='import-form'>"
    html += f"<form id='uploadForm_permissions' action='/admin/import/permissions' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_permissions()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("permissions")
    html += "<form method='post' action='/admin/permissions/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllPermissions'> Выделить всё</label>"
    html += "<ul>"
    for p in permissions:
        html += f"<li><input type='checkbox' name='ids' value='{p.id}'> "
        html += f"{p.permission_name} (код: {p.permission_code})"
        if p.description:
            html += f" – {p.description}"
        html += f" <a href='/admin/permissions/edit/{p.id}'>Редактировать</a> "
        html += f"<a href='/admin/permissions/delete/{p.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<form method='post' action='/admin/permissions/add' style='margin-top: 20px;'>"
    html += "<input type='text' name='permission_code' placeholder='Код права' required>"
    html += "<input type='text' name='permission_name' placeholder='Название' required>"
    html += "<input type='text' name='description' placeholder='Описание'>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllPermissions');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_time_slots_html(time_slots):
    html = f"<h1>Временные слоты (пары)</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/time_slots' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_time_slots()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_time_slots' class='import-form'>"
    html += f"<form id='uploadForm_time_slots' action='/admin/import/time_slots' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_time_slots()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("time_slots")
    html += "<form method='post' action='/admin/time-slots/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllTimeSlots'> Выделить всё</label>"
    html += "</form>"
    html += "<form method='post' action='/admin/time-slots/add' style='margin-top: 20px;'>"
    html += "<input type='number' name='slot_number' placeholder='Номер пары' required>"
    html += "<input type='text' name='name' placeholder='Название (например, 1 пара)' required>"
    html += "<input type='time' name='start_time' placeholder='Время начала (HH:MM)' required>"
    html += "<input type='time' name='end_time' placeholder='Время окончания (HH:MM)' required>"
    html += "<input type='number' name='duration_minutes' placeholder='Длительность (мин)'>"
    html += "<input type='number' name='break_after_minutes' placeholder='Перерыв после (мин)'>"
    html += "<label><input type='checkbox' name='is_active' value='true' checked> Активен</label>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<ul>"
    for ts in time_slots:
        html += f"<li><input type='checkbox' name='ids' value='{ts.id}'> "
        html += f"{ts.name} (№{ts.slot_number}, {ts.start_time}–{ts.end_time})"
        if ts.duration_minutes:
            html += f", длительность: {ts.duration_minutes} мин"
        if not ts.is_active:
            html += " [неактивен]"
        html += f" <a href='/admin/time-slots/edit/{ts.id}'>Редактировать</a> "
        html += f"<a href='/admin/time-slots/delete/{ts.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllTimeSlots');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_subjects_html(subjects):
    html = f"<h1>Дисциплины</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/subjects' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_subjects()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_subjects' class='import-form'>"
    html += f"<form id='uploadForm_subjects' action='/admin/import/subjects' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_subjects()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("subjects")
    html += "<form method='post' action='/admin/subjects/bulk-delete' onsubmit='return confirmDeleteSelected();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllSubjects'> Выделить всё</label>"
    html += "</form>"
    html += "<form method='post' action='/admin/subjects/add' style='margin-top: 20px;'>"
    html += "<input type='text' name='name' placeholder='Название дисциплины' required>"
    html += "<button type='submit'>Добавить</button>"
    html += "</form>"
    html += "<ul>"
    for s in subjects:
        html += f"<li><input type='checkbox' name='ids' value='{s.id}'> "
        html += f"{s.name} "
        html += f"<a href='/admin/subjects/edit/{s.id}'>Редактировать</a> "
        html += f"<a href='/admin/subjects/delete/{s.id}'>Удалить</a></li>"
    html += "</ul>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAll = document.getElementById('selectAllSubjects');
        if (selectAll) {
            selectAll.addEventListener('change', function() {
                const checkboxes = document.querySelectorAll('input[name="ids"]');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
            });
        }
        function confirmDeleteSelected() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Вы уверены, что хотите удалить выбранные записи?');
        }
    </script>
    """
    return html

def render_teacher_groups_html(teacher_groups, teachers, groups):
    html = f"<h1>Связи преподавателей и групп</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/teacher_groups' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_teacher_groups()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_teacher_groups' class='import-form'>"
    html += f"<form id='uploadForm_teacher_groups' action='/admin/import/teacher_groups' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_teacher_groups()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("teacher_groups")
    # Форма добавления с чекбоксами
    html += "<form method='post' action='/admin/teacher-groups/add'>"
    html += "<div style='display: flex; gap: 20px;'>"
    html += "<div><label>Преподаватели:</label><br>"
    html += "<div style='max-height: 200px; overflow-y: auto; border:1px solid #ccc; padding:8px; width: 250px;'>"
    for t in teachers:
        html += f"<label><input type='checkbox' name='teacher_ids' value='{t.id}'> {t.name}</label><br>"
    html += "</div></div>"
    html += "<div><label>Группы:</label><br>"
    html += "<div style='max-height: 200px; overflow-y: auto; border:1px solid #ccc; padding:8px; width: 250px;'>"
    for g in groups:
        html += f"<label><input type='checkbox' name='group_ids' value='{g.id}'> {g.name}</label><br>"
    html += "</div></div>"
    html += "</div>"
    html += "<button type='submit' style='margin-top: 10px;'>Добавить выбранные комбинации</button>"
    html += "</form>"
    # Форма удаления
    html += "<form method='post' action='/admin/teacher-groups/bulk-delete' onsubmit='return confirmDeleteSelectedTG();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllTG'> Выделить всё</label>"
    html += "<ul>"
    for tg in teacher_groups:
        html += f"<li><input type='checkbox' name='ids' value='{tg['teacher_id']},{tg['group_id']}'> "
        html += f"{tg['teacher_id']} – {tg['group_id']} (Преподаватель ID: {tg['teacher_id']}, Группа ID: {tg['group_id']}) "
        html += f"<a href='/admin/teacher-groups/delete/{tg['teacher_id']}/{tg['group_id']}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAllTG = document.getElementById('selectAllTG');
        if(selectAllTG) selectAllTG.addEventListener('change', function() {
            document.querySelectorAll('input[name="ids"]').forEach(cb => cb.checked = selectAllTG.checked);
        });
        function confirmDeleteSelectedTG() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Удалить выбранные связи?');
        }
    </script>
    """
    return html

def render_teacher_subjects_html(teacher_subjects, teachers, subjects):
    html = f"<h1>Связи преподавателей и предметов</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/teacher_subjects' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_teacher_subjects()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_teacher_subjects' class='import-form'>"
    html += f"<form id='uploadForm_teacher_subjects' action='/admin/import/teacher_subjects' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_teacher_subjects()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("teacher_subjects")
    # Форма добавления
    html += "<form method='post' action='/admin/teacher-subjects/add'>"
    html += "<div style='display: flex; gap: 20px; margin-bottom: 10px;'>"
    html += "<div><label>Преподаватели:</label><br>"
    html += "<div style='max-height: 200px; overflow-y: auto; border:1px solid #ccc; padding:8px; width: 250px;'>"
    for t in teachers:
        html += f"<label><input type='checkbox' name='teacher_ids' value='{t.id}'> {t.name}</label><br>"
    html += "</div></div>"
    html += "<div><label>Предметы:</label><br>"
    html += "<div style='max-height: 200px; overflow-y: auto; border:1px solid #ccc; padding:8px; width: 250px;'>"
    for s in subjects:
        html += f"<label><input type='checkbox' name='subject_ids' value='{s.id}'> {s.name}</label><br>"
    html += "</div></div>"
    html += "</div>"
    html += "<label><input type='checkbox' name='is_main' value='true'> Основной преподаватель (для всех выбранных комбинаций)</label><br>"
    html += "<button type='submit'>Добавить выбранные комбинации</button>"
    html += "</form>"
    # Форма удаления
    html += "<form method='post' action='/admin/teacher-subjects/bulk-delete' onsubmit='return confirmDeleteSelectedTS();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllTS'> Выделить всё</label>"
    html += "<ul>"
    for ts in teacher_subjects:
        html += f"<li><input type='checkbox' name='ids' value='{ts['teacher_id']},{ts['subject_id']}'> "
        html += f"Преподаватель ID: {ts['teacher_id']}, Предмет ID: {ts['subject_id']}, Основной: {ts['is_main']} "
        html += f"<a href='/admin/teacher-subjects/edit/{ts['teacher_id']}/{ts['subject_id']}'>Редактировать</a> "
        html += f"<a href='/admin/teacher-subjects/delete/{ts['teacher_id']}/{ts['subject_id']}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAllTS = document.getElementById('selectAllTS');
        if(selectAllTS) selectAllTS.addEventListener('change', function() {
            document.querySelectorAll('input[name="ids"]').forEach(cb => cb.checked = selectAllTS.checked);
        });
        function confirmDeleteSelectedTS() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Удалить выбранные связи?');
        }
    </script>
    """
    return html

def render_audience_subjects_html(audience_subjects, audiences, subjects):
    html = f"<h1>Связи аудиторий и предметов</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/audience_subjects' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_audience_subjects()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_audience_subjects' class='import-form'>"
    html += f"<form id='uploadForm_audience_subjects' action='/admin/import/audience_subjects' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_audience_subjects()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("audience_subjects")
    # Форма добавления
    html += "<form method='post' action='/admin/audience-subjects/add'>"
    html += "<div style='display: flex; gap: 20px; margin-bottom: 10px;'>"
    html += "<div><label>Аудитории:</label><br>"
    html += "<div style='max-height: 200px; overflow-y: auto; border:1px solid #ccc; padding:8px; width: 250px;'>"
    for a in audiences:
        html += f"<label><input type='checkbox' name='audience_ids' value='{a.id}'> {a.name}</label><br>"
    html += "</div></div>"
    html += "<div><label>Предметы:</label><br>"
    html += "<div style='max-height: 200px; overflow-y: auto; border:1px solid #ccc; padding:8px; width: 250px;'>"
    for s in subjects:
        html += f"<label><input type='checkbox' name='subject_ids' value='{s.id}'> {s.name}</label><br>"
    html += "</div></div>"
    html += "</div>"
    html += "<button type='submit'>Добавить выбранные комбинации</button>"
    html += "</form>"
    # Форма удаления
    html += "<form method='post' action='/admin/audience-subjects/bulk-delete' onsubmit='return confirmDeleteSelectedAS();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllAS'> Выделить всё</label>"
    html += "<ul>"
    for asub in audience_subjects:
        html += f"<li><input type='checkbox' name='ids' value='{asub['audience_id']},{asub['subject_id']}'> "
        html += f"Аудитория ID: {asub['audience_id']}, Предмет ID: {asub['subject_id']} "
        html += f"<a href='/admin/audience-subjects/delete/{asub['audience_id']}/{asub['subject_id']}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAllAS = document.getElementById('selectAllAS');
        if(selectAllAS) selectAllAS.addEventListener('change', function() {
            document.querySelectorAll('input[name="ids"]').forEach(cb => cb.checked = selectAllAS.checked);
        });
        function confirmDeleteSelectedAS() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Удалить выбранные связи?');
        }
    </script>
    """
    return html

def render_group_subjects_html(group_subjects, groups, subjects):
    html = f"<h1>Связи групп и предметов</h1>{_common_styles()}"
    html += "<div><a href='/admin/export/group_subjects' class='btn-excel'>📎 Экспорт в Excel</a>"
    html += " <button onclick='toggleImportForm_group_subjects()' class='btn-excel btn-import'>📂 Импорт из Excel</button></div>"
    html += f"<div id='importForm_group_subjects' class='import-form'>"
    html += f"<form id='uploadForm_group_subjects' action='/admin/import/group_subjects' method='post' enctype='multipart/form-data'>"
    html += "<input type='file' name='file' accept='.xlsx,.xls' required>"
    html += "<button type='submit'>Загрузить</button>"
    html += "<button type='button' onclick='toggleImportForm_group_subjects()'>Отмена</button>"
    html += "</form></div>"
    html += _import_script("group_subjects")
    # Форма добавления
    html += "<form method='post' action='/admin/group-subjects/add'>"
    html += "<div style='display: flex; gap: 20px; margin-bottom: 10px;'>"
    html += "<div><label>Группы:</label><br>"
    html += "<div style='max-height: 200px; overflow-y: auto; border:1px solid #ccc; padding:8px; width: 250px;'>"
    for g in groups:
        html += f"<label><input type='checkbox' name='group_ids' value='{g.id}'> {g.name}</label><br>"
    html += "</div></div>"
    html += "<div><label>Предметы:</label><br>"
    html += "<div style='max-height: 200px; overflow-y: auto; border:1px solid #ccc; padding:8px; width: 250px;'>"
    for s in subjects:
        html += f"<label><input type='checkbox' name='subject_ids' value='{s.id}'> {s.name}</label><br>"
    html += "</div></div>"
    html += "</div>"
    html += "<button type='submit'>Добавить выбранные комбинации</button>"
    html += "</form>"
    # Форма удаления
    html += "<form method='post' action='/admin/group-subjects/bulk-delete' onsubmit='return confirmDeleteSelectedGS();'>"
    html += "<button type='submit'>Удалить выбранные</button>"
    html += "<label><input type='checkbox' id='selectAllGS'> Выделить всё</label>"
    html += "<ul>"
    for gs in group_subjects:
        html += f"<li><input type='checkbox' name='ids' value='{gs['group_id']},{gs['subject_id']}'> "
        html += f"Группа ID: {gs['group_id']}, Предмет ID: {gs['subject_id']} "
        html += f"<a href='/admin/group-subjects/delete/{gs['group_id']}/{gs['subject_id']}'>Удалить</a></li>"
    html += "</ul>"
    html += "</form>"
    html += "<a href='/admin/'>На главную админки</a>"
    html += """
    <script>
        const selectAllGS = document.getElementById('selectAllGS');
        if(selectAllGS) selectAllGS.addEventListener('change', function() {
            document.querySelectorAll('input[name="ids"]').forEach(cb => cb.checked = selectAllGS.checked);
        });
        function confirmDeleteSelectedGS() {
            const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
            if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
            return confirm('Удалить выбранные связи?');
        }
    </script>
    """
    return html

# ---------- Маршруты для админки ----------
@router.get("/", response_class=HTMLResponse)
async def admin_home():
    return HTMLResponse(content=f"""
    {_common_styles()}
    <h1>Административная панель</h1>
    <ul>
        <li><a href='/admin/manual-schedule'>Ручное составление расписания</a></li>
        <li><a href='/admin/groups'>Группы</a></li>
        <li><a href='/admin/teachers'>Преподаватели</a></li>
        <li><a href='/admin/audiences'>Аудитории</a></li>
        <li><a href='/admin/buildings'>Здания</a></li>
        <li><a href='/admin/calendar'>Календарь</a></li>
        <li><a href='/admin/faculties'>Факультеты</a></li>
        <li><a href='/admin/users'>Пользователи</a></li>
        <li><a href='/admin/user-types'>Типы пользователей</a></li>
        <li><a href='/admin/subdivisions'>Подразделения</a></li>
        <li><a href='/admin/roles'>Роли</a></li>
        <li><a href='/admin/permissions'>Права</a></li>
        <li><a href='/admin/time-slots'>Временные слоты</a></li>
        <li><a href='/admin/subjects'>Дисциплины</a></li>
        <li><a href='/admin/teacher-groups'>Связи: преподаватели-группы</a></li>
        <li><a href='/admin/teacher-subjects'>Связи: преподаватели-предметы</a></li>
        <li><a href='/admin/audience-subjects'>Связи: аудитории-предметы</a></li>
        <li><a href='/admin/group-subjects'>Связи: группы-предметы</a></li>
        <li><a href='/admin/stats'>Статистика</a></li>
        <li><a href='/admin/schedule-editor'>Черновик расписания</a></li>
        <li><a href='/admin/final-schedule'>Опубликованное расписание</a></li>
    </ul>
    <a href='/'>На главную API</a>
    """)

# ----- ГРУППЫ -----
@router.get("/groups", response_class=HTMLResponse)
async def groups_page(db: AsyncSession = Depends(get_db)):
    groups = await group_service.get_all_groups(db)
    faculties = await faculty_service.get_all_faculties(db)
    return HTMLResponse(content=render_groups_html(groups, faculties))

@router.post("/groups/add", response_model=None)
async def add_group(name: str = Form(...), faculty_id: int = Form(...), student_count: int = Form(None), db: AsyncSession = Depends(get_db)):
    data = GroupCreate(name=name, faculty_id=faculty_id, student_count=student_count)
    await group_service.create_group(db, data)
    return RedirectResponse(url="/admin/groups", status_code=303)

@router.get("/groups/edit/{group_id}", response_class=HTMLResponse)
async def edit_group_form(group_id: int, db: AsyncSession = Depends(get_db)):
    group = await group_service.get_group(db, group_id)
    if not group:
        return HTMLResponse("Группа не найдена", status_code=404)
    faculties = await faculty_service.get_all_faculties(db)
    html = f"""
    <h1>Редактировать группу</h1>
    <form method="post" action="/admin/groups/edit/{group_id}">
        <input type="text" name="name" value="{group.name}" required>
        <select name="faculty_id">
            {''.join(f'<option value="{f.id}" {"selected" if f.id == group.faculty_id else ""}>{f.name}</option>' for f in faculties)}
        </select>
        <input type="number" name="student_count" value="{group.student_count or ''}">
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/groups">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/groups/edit/{group_id}", response_model=None)
async def edit_group(group_id: int, name: str = Form(...), faculty_id: int = Form(...), student_count: int = Form(None), db: AsyncSession = Depends(get_db)):
    data = GroupUpdate(name=name, faculty_id=faculty_id, student_count=student_count)
    await group_service.update_group(db, group_id, data)
    return RedirectResponse(url="/admin/groups", status_code=303)

@router.get("/groups/delete/{group_id}", response_class=HTMLResponse)
async def confirm_delete_group(group_id: int, db: AsyncSession = Depends(get_db)):
    group = await group_service.get_group(db, group_id)
    if not group:
        return HTMLResponse("Группа не найдена", status_code=404)
    html = f"""
    <h1>Удалить группу "{group.name}"?</h1>
    <form method="post" action="/admin/groups/delete/{group_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/groups">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/groups/delete/{group_id}", response_model=None)
async def delete_group(group_id: int, db: AsyncSession = Depends(get_db)):
    await group_service.delete_group(db, group_id)
    return RedirectResponse(url="/admin/groups", status_code=303)

@router.post("/groups/bulk-delete", response_model=None)
async def bulk_delete_groups(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await group_service.bulk_delete_groups(db, ids)
    return RedirectResponse(url="/admin/groups", status_code=303)

# ----- ПРЕПОДАВАТЕЛИ -----
@router.get("/teachers", response_class=HTMLResponse)
async def teachers_page(db: AsyncSession = Depends(get_db)):
    teachers = await teacher_service.get_all_teachers(db)
    return HTMLResponse(content=render_teachers_html(teachers))

@router.post("/teachers/add", response_model=None)
async def add_teacher(login: str = Form(...), name: str = Form(...), url: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = TeacherCreate(login=login, name=name, url=url)
    await teacher_service.create_teacher(db, data)
    return RedirectResponse(url="/admin/teachers", status_code=303)

@router.get("/teachers/edit/{teacher_id}", response_class=HTMLResponse)
async def edit_teacher_form(teacher_id: int, db: AsyncSession = Depends(get_db)):
    teacher = await teacher_service.get_teacher(db, teacher_id)
    if not teacher:
        return HTMLResponse("Преподаватель не найден", status_code=404)
    html = f"""
    <h1>Редактировать преподавателя</h1>
    <form method="post" action="/admin/teachers/edit/{teacher_id}">
        <input type="text" name="login" value="{teacher.login}" required>
        <input type="text" name="name" value="{teacher.name}" required>
        <input type="text" name="url" value="{teacher.url or ''}">
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/teachers">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/teachers/edit/{teacher_id}", response_model=None)
async def edit_teacher(teacher_id: int, login: str = Form(...), name: str = Form(...), url: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = TeacherUpdate(login=login, name=name, url=url)
    await teacher_service.update_teacher(db, teacher_id, data)
    return RedirectResponse(url="/admin/teachers", status_code=303)

@router.get("/teachers/delete/{teacher_id}", response_class=HTMLResponse)
async def confirm_delete_teacher(teacher_id: int, db: AsyncSession = Depends(get_db)):
    teacher = await teacher_service.get_teacher(db, teacher_id)
    if not teacher:
        return HTMLResponse("Преподаватель не найден", status_code=404)
    if await has_dependent_teacher_groups(db, teacher_id) or await has_dependent_teacher_subjects(db, teacher_id):
        html = f"""
        <h1>Ошибка удаления преподавателя "{teacher.name}"</h1>
        <p>Невозможно удалить преподавателя, так как он связан с группами или предметами. Сначала удалите эти связи.</p>
        <a href="/admin/teachers">Вернуться к списку преподавателей</a>
        """
        return HTMLResponse(content=html, status_code=400)
    html = f"""
    <h1>Удалить преподавателя "{teacher.name}"?</h1>
    <form method="post" action="/admin/teachers/delete/{teacher_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/teachers">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/teachers/delete/{teacher_id}", response_model=None)
async def delete_teacher(teacher_id: int, db: AsyncSession = Depends(get_db)):
    await teacher_service.delete_teacher(db, teacher_id)
    return RedirectResponse(url="/admin/teachers", status_code=303)

@router.post("/teachers/bulk-delete", response_model=None)
async def bulk_delete_teachers(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    forbidden = []
    for tid in ids:
        if await has_dependent_teacher_groups(db, tid) or await has_dependent_teacher_subjects(db, tid):
            forbidden.append(tid)
    if forbidden:
        html = f"""
        <h1>Ошибка массового удаления</h1>
        <p>Следующие преподаватели имеют связи с группами или предметами и не могут быть удалены: {forbidden}</p>
        <a href="/admin/teachers">Вернуться к списку преподавателей</a>
        """
        return HTMLResponse(content=html, status_code=400)
    await teacher_service.bulk_delete_teachers(db, ids)
    return RedirectResponse(url="/admin/teachers", status_code=303)

# ----- АУДИТОРИИ -----
@router.get("/audiences", response_class=HTMLResponse)
async def audiences_page(db: AsyncSession = Depends(get_db)):
    audiences = await audience_service.get_all_audiences(db)
    buildings = await building_service.get_all_buildings(db)
    return HTMLResponse(content=render_audiences_html(audiences, buildings))

@router.post("/audiences/add", response_model=None)
async def add_audience(name: str = Form(...), building_id: int = Form(None), capacity: int = Form(None), type: str = Form(None), is_active: bool = Form(True), db: AsyncSession = Depends(get_db)):
    data = AudienceCreate(name=name, building_id=building_id, capacity=capacity, type=type, is_active=is_active)
    await audience_service.create_audience(db, data)
    return RedirectResponse(url="/admin/audiences", status_code=303)

@router.get("/audiences/edit/{audience_id}", response_class=HTMLResponse)
async def edit_audience_form(audience_id: int, db: AsyncSession = Depends(get_db)):
    audience = await audience_service.get_audience(db, audience_id)
    if not audience:
        return HTMLResponse("Аудитория не найдена", status_code=404)
    buildings = await building_service.get_all_buildings(db)
    html = f"""
    <h1>Редактировать аудиторию</h1>
    <form method="post" action="/admin/audiences/edit/{audience_id}">
        <input type="text" name="name" value="{audience.name}" required>
        <select name="building_id">
            <option value="">Не выбрано</option>
            {''.join(f'<option value="{b.id}" {"selected" if b.id == audience.building_id else ""}>{b.name}</option>' for b in buildings)}
        </select>
        <input type="number" name="capacity" value="{audience.capacity or ''}">
        <select name="type">
            <option value="">Тип</option>
            <option value="лекционная" {'selected' if audience.type == 'лекционная' else ''}>Лекционная</option>
            <option value="практическая" {'selected' if audience.type == 'практическая' else ''}>Практическая</option>
            <option value="лабораторная" {'selected' if audience.type == 'лабораторная' else ''}>Лабораторная</option>
            <option value="компьютерный класс" {'selected' if audience.type == 'компьютерный класс' else ''}>Компьютерный класс</option>
        </select>
        <label><input type="checkbox" name="is_active" value="true" {'checked' if audience.is_active else ''}> Активна</label>
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/audiences">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/audiences/edit/{audience_id}", response_model=None)
async def edit_audience(audience_id: int, name: str = Form(...), building_id: int = Form(None), capacity: int = Form(None), type: str = Form(None), is_active: bool = Form(True), db: AsyncSession = Depends(get_db)):
    data = AudienceUpdate(name=name, building_id=building_id, capacity=capacity, type=type, is_active=is_active)
    await audience_service.update_audience(db, audience_id, data)
    return RedirectResponse(url="/admin/audiences", status_code=303)

@router.get("/audiences/delete/{audience_id}", response_class=HTMLResponse)
async def confirm_delete_audience(audience_id: int, db: AsyncSession = Depends(get_db)):
    audience = await audience_service.get_audience(db, audience_id)
    if not audience:
        return HTMLResponse("Аудитория не найдена", status_code=404)
    html = f"""
    <h1>Удалить аудиторию "{audience.name}"?</h1>
    <form method="post" action="/admin/audiences/delete/{audience_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/audiences">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/audiences/delete/{audience_id}", response_model=None)
async def delete_audience(audience_id: int, db: AsyncSession = Depends(get_db)):
    await audience_service.delete_audience(db, audience_id)
    return RedirectResponse(url="/admin/audiences", status_code=303)

@router.post("/audiences/bulk-delete", response_model=None)
async def bulk_delete_audiences(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await audience_service.bulk_delete_audiences(db, ids)
    return RedirectResponse(url="/admin/audiences", status_code=303)

# ----- ЗДАНИЯ -----
@router.get("/buildings", response_class=HTMLResponse)
async def buildings_page(db: AsyncSession = Depends(get_db)):
    buildings = await building_service.get_all_buildings(db)
    return HTMLResponse(content=render_buildings_html(buildings))

@router.post("/buildings/add", response_model=None)
async def add_building(name: str = Form(...), address: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = BuildingCreate(name=name, address=address)
    await building_service.create_building(db, data)
    return RedirectResponse(url="/admin/buildings", status_code=303)

@router.get("/buildings/edit/{building_id}", response_class=HTMLResponse)
async def edit_building_form(building_id: int, db: AsyncSession = Depends(get_db)):
    building = await building_service.get_building(db, building_id)
    if not building:
        return HTMLResponse("Здание не найдено", status_code=404)
    html = f"""
    <h1>Редактировать здание</h1>
    <form method="post" action="/admin/buildings/edit/{building_id}">
        <input type="text" name="name" value="{building.name}" required>
        <input type="text" name="address" value="{building.address or ''}">
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/buildings">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/buildings/edit/{building_id}", response_model=None)
async def edit_building(building_id: int, name: str = Form(...), address: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = BuildingUpdate(name=name, address=address)
    await building_service.update_building(db, building_id, data)
    return RedirectResponse(url="/admin/buildings", status_code=303)

@router.get("/buildings/delete/{building_id}", response_class=HTMLResponse)
async def confirm_delete_building(building_id: int, db: AsyncSession = Depends(get_db)):
    building = await building_service.get_building(db, building_id)
    if not building:
        return HTMLResponse("Здание не найдено", status_code=404)
    html = f"""
    <h1>Удалить здание "{building.name}"?</h1>
    <form method="post" action="/admin/buildings/delete/{building_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/buildings">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/buildings/delete/{building_id}", response_model=None)
async def delete_building(building_id: int, db: AsyncSession = Depends(get_db)):
    await building_service.delete_building(db, building_id)
    return RedirectResponse(url="/admin/buildings", status_code=303)

@router.post("/buildings/bulk-delete", response_model=None)
async def bulk_delete_buildings(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await building_service.bulk_delete_buildings(db, ids)
    return RedirectResponse(url="/admin/buildings", status_code=303)

# ----- КАЛЕНДАРЬ -----
@router.get("/calendar", response_class=HTMLResponse)
async def calendar_page(db: AsyncSession = Depends(get_db)):
    entries = await calendar_service.get_all_calendar(db)
    return HTMLResponse(content=render_calendar_html(entries))

@router.post("/calendar/add", response_model=None)
async def add_calendar_entry(date: date = Form(...), is_working_day: bool = Form(True), description: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = CalendarCreate(date=date, is_working_day=is_working_day, description=description)
    await calendar_service.create_calendar_entry(db, data)
    return RedirectResponse(url="/admin/calendar", status_code=303)

@router.post("/calendar/sync", response_model=None)
async def sync_calendar_web(year: int = Form(...), background_tasks: BackgroundTasks = BackgroundTasks()):
    background_tasks.add_task(calendar_sync.sync_calendar_for_year, year)
    return RedirectResponse(url="/admin/calendar", status_code=303)

@router.get("/calendar/edit/{entry_id}", response_class=HTMLResponse)
async def edit_calendar_form(entry_id: int, db: AsyncSession = Depends(get_db)):
    entry = await calendar_service.get_calendar_entry(db, entry_id)
    if not entry:
        return HTMLResponse("Запись не найдена", status_code=404)
    html = f"""
    <h1>Редактировать запись календаря</h1>
    <form method="post" action="/admin/calendar/edit/{entry_id}">
        <input type="date" name="date" value="{entry.date}" required>
        <select name="is_working_day">
            <option value="true" {'selected' if entry.is_working_day else ''}>Рабочий день</option>
            <option value="false" {'selected' if not entry.is_working_day else ''}>Выходной/праздник</option>
        </select>
        <input type="text" name="description" value="{entry.description or ''}">
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/calendar">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/calendar/edit/{entry_id}", response_model=None)
async def edit_calendar_entry(entry_id: int, date: date = Form(...), is_working_day: bool = Form(True), description: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = CalendarUpdate(date=date, is_working_day=is_working_day, description=description)
    await calendar_service.update_calendar_entry(db, entry_id, data)
    return RedirectResponse(url="/admin/calendar", status_code=303)

@router.get("/calendar/delete/{entry_id}", response_class=HTMLResponse)
async def confirm_delete_calendar(entry_id: int, db: AsyncSession = Depends(get_db)):
    entry = await calendar_service.get_calendar_entry(db, entry_id)
    if not entry:
        return HTMLResponse("Запись не найдена", status_code=404)
    html = f"""
    <h1>Удалить запись за {entry.date}?</h1>
    <form method="post" action="/admin/calendar/delete/{entry_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/calendar">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/calendar/delete/{entry_id}", response_model=None)
async def delete_calendar_entry(entry_id: int, db: AsyncSession = Depends(get_db)):
    await calendar_service.delete_calendar_entry(db, entry_id)
    return RedirectResponse(url="/admin/calendar", status_code=303)

@router.post("/calendar/bulk-delete", response_model=None)
async def bulk_delete_calendar(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await calendar_service.bulk_delete_calendar(db, ids)
    return RedirectResponse(url="/admin/calendar", status_code=303)

# ----- ФАКУЛЬТЕТЫ -----
@router.get("/faculties", response_class=HTMLResponse)
async def faculties_page(db: AsyncSession = Depends(get_db)):
    faculties = await faculty_service.get_all_faculties(db)
    return HTMLResponse(content=render_faculties_html(faculties))

@router.post("/faculties/add", response_model=None)
async def add_faculty(name: str = Form(...), display_name: str = Form(None), short_display_name: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = FacultyCreate(name=name, display_name=display_name, short_display_name=short_display_name)
    await faculty_service.create_faculty(db, data)
    return RedirectResponse(url="/admin/faculties", status_code=303)

@router.get("/faculties/edit/{faculty_id}", response_class=HTMLResponse)
async def edit_faculty_form(faculty_id: int, db: AsyncSession = Depends(get_db)):
    faculty = await faculty_service.get_faculty(db, faculty_id)
    if not faculty:
        return HTMLResponse("Факультет не найден", status_code=404)
    html = f"""
    <h1>Редактировать факультет</h1>
    <form method="post" action="/admin/faculties/edit/{faculty_id}">
        <input type="text" name="name" value="{faculty.name}" required>
        <input type="text" name="display_name" value="{faculty.display_name or ''}">
        <input type="text" name="short_display_name" value="{faculty.short_display_name or ''}">
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/faculties">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/faculties/edit/{faculty_id}", response_model=None)
async def edit_faculty(faculty_id: int, name: str = Form(...), display_name: str = Form(None), short_display_name: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = FacultyUpdate(name=name, display_name=display_name, short_display_name=short_display_name)
    await faculty_service.update_faculty(db, faculty_id, data)
    return RedirectResponse(url="/admin/faculties", status_code=303)

@router.get("/faculties/delete/{faculty_id}", response_class=HTMLResponse)
async def confirm_delete_faculty(faculty_id: int, db: AsyncSession = Depends(get_db)):
    faculty = await faculty_service.get_faculty(db, faculty_id)
    if not faculty:
        return HTMLResponse("Факультет не найден", status_code=404)
    if await has_dependent_groups(db, faculty_id):
        html = f"""
        <h1>Ошибка удаления факультета "{faculty.name}"</h1>
        <p>Невозможно удалить факультет, так как к нему привязаны группы. Сначала удалите или переназначьте группы.</p>
        <a href="/admin/faculties">Вернуться к списку факультетов</a>
        """
        return HTMLResponse(content=html, status_code=400)
    html = f"""
    <h1>Удалить факультет "{faculty.name}"?</h1>
    <form method="post" action="/admin/faculties/delete/{faculty_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/faculties">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/faculties/delete/{faculty_id}", response_model=None)
async def delete_faculty(faculty_id: int, db: AsyncSession = Depends(get_db)):
    await faculty_service.delete_faculty(db, faculty_id)
    return RedirectResponse(url="/admin/faculties", status_code=303)

@router.post("/faculties/bulk-delete", response_model=None)
async def bulk_delete_faculties(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    forbidden = []
    for fid in ids:
        if await has_dependent_groups(db, fid):
            forbidden.append(fid)
    if forbidden:
        html = f"""
        <h1>Ошибка массового удаления</h1>
        <p>Следующие факультеты имеют привязанные группы и не могут быть удалены: {forbidden}</p>
        <a href="/admin/faculties">Вернуться к списку факультетов</a>
        """
        return HTMLResponse(content=html, status_code=400)
    await faculty_service.bulk_delete_faculties(db, ids)
    return RedirectResponse(url="/admin/faculties", status_code=303)

# ----- ТИПЫ ПОЛЬЗОВАТЕЛЕЙ -----
@router.get("/user-types", response_class=HTMLResponse)
async def user_types_page(db: AsyncSession = Depends(get_db)):
    user_types = await user_type_service.get_all_user_types(db)
    return HTMLResponse(content=render_user_types_html(user_types))

@router.post("/user-types/add", response_model=None)
async def add_user_type(code: str = Form(...), name: str = Form(...), description: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = UserTypeCreate(code=code, name=name, description=description)
    await user_type_service.create_user_type(db, data)
    return RedirectResponse(url="/admin/user-types", status_code=303)

@router.get("/user-types/edit/{ut_id}", response_class=HTMLResponse)
async def edit_user_type_form(ut_id: int, db: AsyncSession = Depends(get_db)):
    ut = await user_type_service.get_user_type(db, ut_id)
    if not ut:
        return HTMLResponse("Тип пользователя не найден", status_code=404)
    html = f"""
    <h1>Редактировать тип пользователя</h1>
    <form method="post" action="/admin/user-types/edit/{ut_id}">
        <input type="text" name="code" value="{ut.code}" required>
        <input type="text" name="name" value="{ut.name}" required>
        <input type="text" name="description" value="{ut.description or ''}">
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/user-types">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/user-types/edit/{ut_id}", response_model=None)
async def edit_user_type(ut_id: int, code: str = Form(...), name: str = Form(...), description: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = UserTypeUpdate(code=code, name=name, description=description)
    await user_type_service.update_user_type(db, ut_id, data)
    return RedirectResponse(url="/admin/user-types", status_code=303)

@router.get("/user-types/delete/{ut_id}", response_class=HTMLResponse)
async def confirm_delete_user_type(ut_id: int, db: AsyncSession = Depends(get_db)):
    ut = await user_type_service.get_user_type(db, ut_id)
    if not ut:
        return HTMLResponse("Тип пользователя не найден", status_code=404)
    html = f"""
    <h1>Удалить тип "{ut.name}"?</h1>
    <form method="post" action="/admin/user-types/delete/{ut_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/user-types">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/user-types/delete/{ut_id}", response_model=None)
async def delete_user_type(ut_id: int, db: AsyncSession = Depends(get_db)):
    await user_type_service.delete_user_type(db, ut_id)
    return RedirectResponse(url="/admin/user-types", status_code=303)

@router.post("/user-types/bulk-delete", response_model=None)
async def bulk_delete_user_types(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await user_type_service.bulk_delete_user_types(db, ids)
    return RedirectResponse(url="/admin/user-types", status_code=303)

# ----- ПОЛЬЗОВАТЕЛИ -----
@router.get("/users", response_class=HTMLResponse)
async def users_page(db: AsyncSession = Depends(get_db)):
    users = await user_service.get_all_users(db)
    user_types = await user_type_service.get_all_user_types(db)
    return HTMLResponse(content=render_users_html(users, user_types))

@router.post("/users/add", response_model=None)
async def add_user(
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    full_name: str = Form(None),
    user_type_id: int = Form(None),
    is_active: bool = Form(True),
    db: AsyncSession = Depends(get_db)
):
    data = UserCreate(
        username=username,
        email=email,
        password=password,
        full_name=full_name,
        user_type_id=user_type_id,
        is_active=is_active
    )
    await user_service.create_user(db, data)
    return RedirectResponse(url="/admin/users", status_code=303)

@router.get("/users/edit/{user_id}", response_class=HTMLResponse)
async def edit_user_form(user_id: int, db: AsyncSession = Depends(get_db)):
    user = await user_service.get_user(db, user_id)
    if not user:
        return HTMLResponse("Пользователь не найден", status_code=404)
    user_types = await user_type_service.get_all_user_types(db)
    html = f"""
    <h1>Редактировать пользователя {user.username}</h1>
    <form method="post" action="/admin/users/edit/{user_id}">
        <input type="text" name="username" value="{user.username}" required>
        <input type="email" name="email" value="{user.email}" required>
        <input type="password" name="password" placeholder="Новый пароль (оставьте пустым, чтобы не менять)">
        <input type="text" name="full_name" value="{user.full_name or ''}">
        <select name="user_type_id">
            <option value="">Тип пользователя</option>
    """
    for ut in user_types:
        selected = "selected" if ut.id == user.user_type_id else ""
        html += f'<option value="{ut.id}" {selected}>{ut.name}</option>'
    html += f"""
        </select>
        <label><input type="checkbox" name="is_active" value="true" {'checked' if user.is_active else ''}> Активен</label>
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/users">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/users/edit/{user_id}", response_model=None)
async def edit_user(
    user_id: int,
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(None),
    full_name: str = Form(None),
    user_type_id: int = Form(None),
    is_active: bool = Form(True),
    db: AsyncSession = Depends(get_db)
):
    data = UserUpdate(
        username=username,
        email=email,
        password=password if password else None,
        full_name=full_name,
        user_type_id=user_type_id,
        is_active=is_active
    )
    await user_service.update_user(db, user_id, data)
    return RedirectResponse(url="/admin/users", status_code=303)

@router.get("/users/delete/{user_id}", response_class=HTMLResponse)
async def confirm_delete_user(user_id: int, db: AsyncSession = Depends(get_db)):
    user = await user_service.get_user(db, user_id)
    if not user:
        return HTMLResponse("Пользователь не найден", status_code=404)
    html = f"""
    <h1>Удалить пользователя "{user.username}"?</h1>
    <form method="post" action="/admin/users/delete/{user_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/users">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/users/delete/{user_id}", response_model=None)
async def delete_user(user_id: int, db: AsyncSession = Depends(get_db)):
    await user_service.delete_user(db, user_id)
    return RedirectResponse(url="/admin/users", status_code=303)

@router.post("/users/bulk-delete", response_model=None)
async def bulk_delete_users(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await user_service.bulk_delete_users(db, ids)
    return RedirectResponse(url="/admin/users", status_code=303)

# ----- ПОДРАЗДЕЛЕНИЯ -----
@router.get("/subdivisions", response_class=HTMLResponse)
async def subdivisions_page(db: AsyncSession = Depends(get_db)):
    subdivisions = await subdivision_service.get_all_subdivisions(db)
    return HTMLResponse(content=render_subdivisions_html(subdivisions))

@router.post("/subdivisions/add", response_model=None)
async def add_subdivision(name: str = Form(...), db: AsyncSession = Depends(get_db)):
    data = SubdivisionCreate(name=name)
    await subdivision_service.create_subdivision(db, data)
    return RedirectResponse(url="/admin/subdivisions", status_code=303)

@router.get("/subdivisions/edit/{sub_id}", response_class=HTMLResponse)
async def edit_subdivision_form(sub_id: int, db: AsyncSession = Depends(get_db)):
    sub = await subdivision_service.get_subdivision(db, sub_id)
    if not sub:
        return HTMLResponse("Подразделение не найдено", status_code=404)
    html = f"""
    <h1>Редактировать подразделение</h1>
    <form method="post" action="/admin/subdivisions/edit/{sub_id}">
        <input type="text" name="name" value="{sub.name}" required>
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/subdivisions">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/subdivisions/edit/{sub_id}", response_model=None)
async def edit_subdivision(sub_id: int, name: str = Form(...), db: AsyncSession = Depends(get_db)):
    data = SubdivisionUpdate(name=name)
    await subdivision_service.update_subdivision(db, sub_id, data)
    return RedirectResponse(url="/admin/subdivisions", status_code=303)

@router.get("/subdivisions/delete/{sub_id}", response_class=HTMLResponse)
async def confirm_delete_subdivision(sub_id: int, db: AsyncSession = Depends(get_db)):
    sub = await subdivision_service.get_subdivision(db, sub_id)
    if not sub:
        return HTMLResponse("Подразделение не найдено", status_code=404)
    html = f"""
    <h1>Удалить подразделение "{sub.name}"?</h1>
    <form method="post" action="/admin/subdivisions/delete/{sub_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/subdivisions">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/subdivisions/delete/{sub_id}", response_model=None)
async def delete_subdivision(sub_id: int, db: AsyncSession = Depends(get_db)):
    await subdivision_service.delete_subdivision(db, sub_id)
    return RedirectResponse(url="/admin/subdivisions", status_code=303)

@router.post("/subdivisions/bulk-delete", response_model=None)
async def bulk_delete_subdivisions(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await subdivision_service.bulk_delete_subdivisions(db, ids)
    return RedirectResponse(url="/admin/subdivisions", status_code=303)

# ----- РОЛИ -----
@router.get("/roles", response_class=HTMLResponse)
async def roles_page(db: AsyncSession = Depends(get_db)):
    roles = await role_service.get_all_roles(db)
    return HTMLResponse(content=render_roles_html(roles))

@router.post("/roles/add", response_model=None)
async def add_role(role_name: str = Form(...), description: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = RoleCreate(role_name=role_name, description=description)
    await role_service.create_role(db, data)
    return RedirectResponse(url="/admin/roles", status_code=303)

@router.get("/roles/edit/{role_id}", response_class=HTMLResponse)
async def edit_role_form(role_id: int, db: AsyncSession = Depends(get_db)):
    role = await role_service.get_role(db, role_id)
    if not role:
        return HTMLResponse("Роль не найдена", status_code=404)
    html = f"""
    <h1>Редактировать роль</h1>
    <form method="post" action="/admin/roles/edit/{role_id}">
        <input type="text" name="role_name" value="{role.role_name}" required>
        <input type="text" name="description" value="{role.description or ''}">
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/roles">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/roles/edit/{role_id}", response_model=None)
async def edit_role(role_id: int, role_name: str = Form(...), description: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = RoleUpdate(role_name=role_name, description=description)
    await role_service.update_role(db, role_id, data)
    return RedirectResponse(url="/admin/roles", status_code=303)

@router.get("/roles/delete/{role_id}", response_class=HTMLResponse)
async def confirm_delete_role(role_id: int, db: AsyncSession = Depends(get_db)):
    role = await role_service.get_role(db, role_id)
    if not role:
        return HTMLResponse("Роль не найдена", status_code=404)
    html = f"""
    <h1>Удалить роль "{role.role_name}"?</h1>
    <form method="post" action="/admin/roles/delete/{role_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/roles">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/roles/delete/{role_id}", response_model=None)
async def delete_role(role_id: int, db: AsyncSession = Depends(get_db)):
    await role_service.delete_role(db, role_id)
    return RedirectResponse(url="/admin/roles", status_code=303)

@router.post("/roles/bulk-delete", response_model=None)
async def bulk_delete_roles(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await role_service.bulk_delete_roles(db, ids)
    return RedirectResponse(url="/admin/roles", status_code=303)

# ----- ПРАВА -----
@router.get("/permissions", response_class=HTMLResponse)
async def permissions_page(db: AsyncSession = Depends(get_db)):
    permissions = await permission_service.get_all_permissions(db)
    return HTMLResponse(content=render_permissions_html(permissions))

@router.post("/permissions/add", response_model=None)
async def add_permission(permission_code: str = Form(...), permission_name: str = Form(...), description: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = PermissionCreate(permission_code=permission_code, permission_name=permission_name, description=description)
    await permission_service.create_permission(db, data)
    return RedirectResponse(url="/admin/permissions", status_code=303)

@router.get("/permissions/edit/{perm_id}", response_class=HTMLResponse)
async def edit_permission_form(perm_id: int, db: AsyncSession = Depends(get_db)):
    perm = await permission_service.get_permission(db, perm_id)
    if not perm:
        return HTMLResponse("Право не найдено", status_code=404)
    html = f"""
    <h1>Редактировать право</h1>
    <form method="post" action="/admin/permissions/edit/{perm_id}">
        <input type="text" name="permission_code" value="{perm.permission_code}" required>
        <input type="text" name="permission_name" value="{perm.permission_name}" required>
        <input type="text" name="description" value="{perm.description or ''}">
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/permissions">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/permissions/edit/{perm_id}", response_model=None)
async def edit_permission(perm_id: int, permission_code: str = Form(...), permission_name: str = Form(...), description: str = Form(None), db: AsyncSession = Depends(get_db)):
    data = PermissionUpdate(permission_code=permission_code, permission_name=permission_name, description=description)
    await permission_service.update_permission(db, perm_id, data)
    return RedirectResponse(url="/admin/permissions", status_code=303)

@router.get("/permissions/delete/{perm_id}", response_class=HTMLResponse)
async def confirm_delete_permission(perm_id: int, db: AsyncSession = Depends(get_db)):
    perm = await permission_service.get_permission(db, perm_id)
    if not perm:
        return HTMLResponse("Право не найдено", status_code=404)
    html = f"""
    <h1>Удалить право "{perm.permission_name}"?</h1>
    <form method="post" action="/admin/permissions/delete/{perm_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/permissions">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/permissions/delete/{perm_id}", response_model=None)
async def delete_permission(perm_id: int, db: AsyncSession = Depends(get_db)):
    await permission_service.delete_permission(db, perm_id)
    return RedirectResponse(url="/admin/permissions", status_code=303)

@router.post("/permissions/bulk-delete", response_model=None)
async def bulk_delete_permissions(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await permission_service.bulk_delete_permissions(db, ids)
    return RedirectResponse(url="/admin/permissions", status_code=303)

# ----- ВРЕМЕННЫЕ СЛОТЫ -----
@router.get("/time-slots", response_class=HTMLResponse)
async def time_slots_page(db: AsyncSession = Depends(get_db)):
    time_slots = await time_slot_service.get_all_time_slots(db)
    return HTMLResponse(content=render_time_slots_html(time_slots))

@router.post("/time-slots/add", response_model=None)
async def add_time_slot(slot_number: int = Form(...), name: str = Form(...), start_time: str = Form(...), end_time: str = Form(...), duration_minutes: int = Form(None), break_after_minutes: int = Form(None), is_active: bool = Form(True), db: AsyncSession = Depends(get_db)):
    from datetime import time
    start = time.fromisoformat(start_time)
    end = time.fromisoformat(end_time)
    data = TimeSlotCreate(slot_number=slot_number, name=name, start_time=start, end_time=end, duration_minutes=duration_minutes, break_after_minutes=break_after_minutes, is_active=is_active)
    await time_slot_service.create_time_slot(db, data)
    return RedirectResponse(url="/admin/time-slots", status_code=303)

@router.get("/time-slots/edit/{ts_id}", response_class=HTMLResponse)
async def edit_time_slot_form(ts_id: int, db: AsyncSession = Depends(get_db)):
    ts = await time_slot_service.get_time_slot(db, ts_id)
    if not ts:
        return HTMLResponse("Слот не найден", status_code=404)
    html = f"""
    <h1>Редактировать временной слот</h1>
    <form method="post" action="/admin/time-slots/edit/{ts_id}">
        <input type="number" name="slot_number" value="{ts.slot_number}" required>
        <input type="text" name="name" value="{ts.name}" required>
        <input type="time" name="start_time" value="{ts.start_time.isoformat()}" required>
        <input type="time" name="end_time" value="{ts.end_time.isoformat()}" required>
        <input type="number" name="duration_minutes" value="{ts.duration_minutes or ''}">
        <input type="number" name="break_after_minutes" value="{ts.break_after_minutes or ''}">
        <label><input type="checkbox" name="is_active" value="true" {'checked' if ts.is_active else ''}> Активен</label>
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/time-slots">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/time-slots/edit/{ts_id}", response_model=None)
async def edit_time_slot(ts_id: int, slot_number: int = Form(...), name: str = Form(...), start_time: str = Form(...), end_time: str = Form(...), duration_minutes: int = Form(None), break_after_minutes: int = Form(None), is_active: bool = Form(True), db: AsyncSession = Depends(get_db)):
    from datetime import time
    start = time.fromisoformat(start_time)
    end = time.fromisoformat(end_time)
    data = TimeSlotUpdate(slot_number=slot_number, name=name, start_time=start, end_time=end, duration_minutes=duration_minutes, break_after_minutes=break_after_minutes, is_active=is_active)
    await time_slot_service.update_time_slot(db, ts_id, data)
    return RedirectResponse(url="/admin/time-slots", status_code=303)

@router.get("/time-slots/delete/{ts_id}", response_class=HTMLResponse)
async def confirm_delete_time_slot(ts_id: int, db: AsyncSession = Depends(get_db)):
    ts = await time_slot_service.get_time_slot(db, ts_id)
    if not ts:
        return HTMLResponse("Слот не найден", status_code=404)
    html = f"""
    <h1>Удалить слот "{ts.name}"?</h1>
    <form method="post" action="/admin/time-slots/delete/{ts_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/time-slots">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/time-slots/delete/{ts_id}", response_model=None)
async def delete_time_slot(ts_id: int, db: AsyncSession = Depends(get_db)):
    await time_slot_service.delete_time_slot(db, ts_id)
    return RedirectResponse(url="/admin/time-slots", status_code=303)

@router.post("/time-slots/bulk-delete", response_model=None)
async def bulk_delete_time_slots(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await time_slot_service.bulk_delete_time_slots(db, ids)
    return RedirectResponse(url="/admin/time-slots", status_code=303)

# ----- ДИСЦИПЛИНЫ -----
@router.get("/subjects", response_class=HTMLResponse)
async def subjects_page(db: AsyncSession = Depends(get_db)):
    subjects = await subject_service.get_all_subjects(db)
    return HTMLResponse(content=render_subjects_html(subjects))

@router.post("/subjects/add", response_model=None)
async def add_subject(name: str = Form(...), db: AsyncSession = Depends(get_db)):
    data = SubjectCreate(name=name)
    await subject_service.create_subject(db, data)
    return RedirectResponse(url="/admin/subjects", status_code=303)

@router.get("/subjects/edit/{subject_id}", response_class=HTMLResponse)
async def edit_subject_form(subject_id: int, db: AsyncSession = Depends(get_db)):
    subject = await subject_service.get_subject(db, subject_id)
    if not subject:
        return HTMLResponse("Дисциплина не найдена", status_code=404)
    html = f"""
    <h1>Редактировать дисциплину</h1>
    <form method="post" action="/admin/subjects/edit/{subject_id}">
        <input type="text" name="name" value="{subject.name}" required>
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/subjects">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/subjects/edit/{subject_id}", response_model=None)
async def edit_subject(subject_id: int, name: str = Form(...), db: AsyncSession = Depends(get_db)):
    data = SubjectUpdate(name=name)
    await subject_service.update_subject(db, subject_id, data)
    return RedirectResponse(url="/admin/subjects", status_code=303)

@router.get("/subjects/delete/{subject_id}", response_class=HTMLResponse)
async def confirm_delete_subject(subject_id: int, db: AsyncSession = Depends(get_db)):
    subject = await subject_service.get_subject(db, subject_id)
    if not subject:
        return HTMLResponse("Дисциплина не найдена", status_code=404)
    html = f"""
    <h1>Удалить дисциплину "{subject.name}"?</h1>
    <form method="post" action="/admin/subjects/delete/{subject_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/subjects">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/subjects/delete/{subject_id}", response_model=None)
async def delete_subject(subject_id: int, db: AsyncSession = Depends(get_db)):
    await subject_service.delete_subject(db, subject_id)
    return RedirectResponse(url="/admin/subjects", status_code=303)

@router.post("/subjects/bulk-delete", response_model=None)
async def bulk_delete_subjects(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    await subject_service.bulk_delete_subjects(db, ids)
    return RedirectResponse(url="/admin/subjects", status_code=303)

# ----- СВЯЗЬ ПРЕПОДАВАТЕЛЬ-ГРУППА -----
@router.get("/teacher-groups", response_class=HTMLResponse)
async def teacher_groups_page(db: AsyncSession = Depends(get_db)):
    teachers = await teacher_service.get_all_teachers(db)
    groups = await group_service.get_all_groups(db)
    teacher_groups = await teacher_group_service.get_all_teacher_groups(db)
    return HTMLResponse(content=render_teacher_groups_html(teacher_groups, teachers, groups))

@router.post("/teacher-groups/add", response_model=None)
async def add_teacher_group(teacher_ids: List[int] = Form(...), group_ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    for t_id in teacher_ids:
        for g_id in group_ids:
            await teacher_group_service.create_teacher_group(db, t_id, g_id)
    return RedirectResponse(url="/admin/teacher-groups", status_code=303)

@router.post("/teacher-groups/bulk-delete", response_model=None)
async def bulk_delete_teacher_groups(ids: List[str] = Form(...), db: AsyncSession = Depends(get_db)):
    pairs = []
    for item in ids:
        parts = item.split(',')
        if len(parts) == 2:
            pairs.append((int(parts[0]), int(parts[1])))
    await teacher_group_service.bulk_delete_teacher_groups(db, pairs)
    return RedirectResponse(url="/admin/teacher-groups", status_code=303)

@router.get("/teacher-groups/delete/{teacher_id}/{group_id}", response_class=HTMLResponse)
async def confirm_delete_teacher_group(teacher_id: int, group_id: int, db: AsyncSession = Depends(get_db)):
    html = f"""
    <h1>Удалить связь преподавателя {teacher_id} и группы {group_id}?</h1>
    <form method="post" action="/admin/teacher-groups/delete/{teacher_id}/{group_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/teacher-groups">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/teacher-groups/delete/{teacher_id}/{group_id}", response_model=None)
async def delete_teacher_group(teacher_id: int, group_id: int, db: AsyncSession = Depends(get_db)):
    await teacher_group_service.delete_teacher_group(db, teacher_id, group_id)
    return RedirectResponse(url="/admin/teacher-groups", status_code=303)

# ----- СВЯЗЬ ПРЕПОДАВАТЕЛЬ-ПРЕДМЕТ -----
@router.get("/teacher-subjects", response_class=HTMLResponse)
async def teacher_subjects_page(db: AsyncSession = Depends(get_db)):
    teachers = await teacher_service.get_all_teachers(db)
    subjects = await subject_service.get_all_subjects(db)
    teacher_subjects = await teacher_subject_service.get_all_teacher_subjects(db)
    return HTMLResponse(content=render_teacher_subjects_html(teacher_subjects, teachers, subjects))

@router.post("/teacher-subjects/add", response_model=None)
async def add_teacher_subject(teacher_ids: List[int] = Form(...), subject_ids: List[int] = Form(...), is_main: bool = Form(False), db: AsyncSession = Depends(get_db)):
    for t_id in teacher_ids:
        for s_id in subject_ids:
            await teacher_subject_service.create_teacher_subject(db, t_id, s_id, is_main)
    return RedirectResponse(url="/admin/teacher-subjects", status_code=303)

@router.post("/teacher-subjects/bulk-delete", response_model=None)
async def bulk_delete_teacher_subjects(ids: List[str] = Form(...), db: AsyncSession = Depends(get_db)):
    pairs = []
    for item in ids:
        parts = item.split(',')
        if len(parts) == 2:
            pairs.append((int(parts[0]), int(parts[1])))
    await teacher_subject_service.bulk_delete_teacher_subjects(db, pairs)
    return RedirectResponse(url="/admin/teacher-subjects", status_code=303)

@router.get("/teacher-subjects/edit/{teacher_id}/{subject_id}", response_class=HTMLResponse)
async def edit_teacher_subject_form(teacher_id: int, subject_id: int, db: AsyncSession = Depends(get_db)):
    from sqlalchemy import select
    from app.models.reference import op_teachers_of_pairs
    result = await db.execute(select(op_teachers_of_pairs).where(op_teachers_of_pairs.c.teacher_id == teacher_id, op_teachers_of_pairs.c.subject_id == subject_id))
    ts = result.first()
    if not ts:
        return HTMLResponse("Связь не найдена", status_code=404)
    is_main_checked = "checked" if ts.is_main else ""
    html = f"""
    <h1>Редактировать связь преподавателя {teacher_id} и предмета {subject_id}</h1>
    <form method="post" action="/admin/teacher-subjects/edit/{teacher_id}/{subject_id}">
        <label><input type="checkbox" name="is_main" value="true" {is_main_checked}> Основной преподаватель</label>
        <button type="submit">Сохранить</button>
    </form>
    <a href="/admin/teacher-subjects">Отмена</a>
    """
    return HTMLResponse(content=html)

@router.post("/teacher-subjects/edit/{teacher_id}/{subject_id}", response_model=None)
async def edit_teacher_subject(teacher_id: int, subject_id: int, is_main: bool = Form(False), db: AsyncSession = Depends(get_db)):
    await teacher_subject_service.update_teacher_subject(db, teacher_id, subject_id, is_main)
    return RedirectResponse(url="/admin/teacher-subjects", status_code=303)

@router.get("/teacher-subjects/delete/{teacher_id}/{subject_id}", response_class=HTMLResponse)
async def confirm_delete_teacher_subject(teacher_id: int, subject_id: int, db: AsyncSession = Depends(get_db)):
    html = f"""
    <h1>Удалить связь преподавателя {teacher_id} и предмета {subject_id}?</h1>
    <form method="post" action="/admin/teacher-subjects/delete/{teacher_id}/{subject_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/teacher-subjects">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/teacher-subjects/delete/{teacher_id}/{subject_id}", response_model=None)
async def delete_teacher_subject(teacher_id: int, subject_id: int, db: AsyncSession = Depends(get_db)):
    await teacher_subject_service.delete_teacher_subject(db, teacher_id, subject_id)
    return RedirectResponse(url="/admin/teacher-subjects", status_code=303)

# ----- СВЯЗЬ АУДИТОРИЯ-ПРЕДМЕТ -----
@router.get("/audience-subjects", response_class=HTMLResponse)
async def audience_subjects_page(db: AsyncSession = Depends(get_db)):
    audiences = await audience_service.get_all_audiences(db)
    subjects = await subject_service.get_all_subjects(db)
    audience_subjects = await audience_subject_service.get_all_audience_subjects(db)
    return HTMLResponse(content=render_audience_subjects_html(audience_subjects, audiences, subjects))

@router.post("/audience-subjects/add", response_model=None)
async def add_audience_subject(audience_ids: List[int] = Form(...), subject_ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    for a_id in audience_ids:
        for s_id in subject_ids:
            await audience_subject_service.create_audience_subject(db, a_id, s_id)
    return RedirectResponse(url="/admin/audience-subjects", status_code=303)

@router.post("/audience-subjects/bulk-delete", response_model=None)
async def bulk_delete_audience_subjects(ids: List[str] = Form(...), db: AsyncSession = Depends(get_db)):
    pairs = []
    for item in ids:
        parts = item.split(',')
        if len(parts) == 2:
            pairs.append((int(parts[0]), int(parts[1])))
    await audience_subject_service.bulk_delete_audience_subjects(db, pairs)
    return RedirectResponse(url="/admin/audience-subjects", status_code=303)

@router.get("/audience-subjects/delete/{audience_id}/{subject_id}", response_class=HTMLResponse)
async def confirm_delete_audience_subject(audience_id: int, subject_id: int, db: AsyncSession = Depends(get_db)):
    html = f"""
    <h1>Удалить связь аудитории {audience_id} и предмета {subject_id}?</h1>
    <form method="post" action="/admin/audience-subjects/delete/{audience_id}/{subject_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/audience-subjects">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/audience-subjects/delete/{audience_id}/{subject_id}", response_model=None)
async def delete_audience_subject(audience_id: int, subject_id: int, db: AsyncSession = Depends(get_db)):
    await audience_subject_service.delete_audience_subject(db, audience_id, subject_id)
    return RedirectResponse(url="/admin/audience-subjects", status_code=303)

# ----- СВЯЗЬ ГРУППА-ПРЕДМЕТ -----
@router.get("/group-subjects", response_class=HTMLResponse)
async def group_subjects_page(db: AsyncSession = Depends(get_db)):
    groups = await group_service.get_all_groups(db)
    subjects = await subject_service.get_all_subjects(db)
    group_subjects = await group_subject_service.get_all_group_subjects(db)
    return HTMLResponse(content=render_group_subjects_html(group_subjects, groups, subjects))

@router.post("/group-subjects/add", response_model=None)
async def add_group_subject(group_ids: List[int] = Form(...), subject_ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    for g_id in group_ids:
        for s_id in subject_ids:
            await group_subject_service.create_group_subject(db, g_id, s_id)
    return RedirectResponse(url="/admin/group-subjects", status_code=303)

@router.post("/group-subjects/bulk-delete", response_model=None)
async def bulk_delete_group_subjects(ids: List[str] = Form(...), db: AsyncSession = Depends(get_db)):
    pairs = []
    for item in ids:
        parts = item.split(',')
        if len(parts) == 2:
            pairs.append((int(parts[0]), int(parts[1])))
    await group_subject_service.bulk_delete_group_subjects(db, pairs)
    return RedirectResponse(url="/admin/group-subjects", status_code=303)

@router.get("/group-subjects/delete/{group_id}/{subject_id}", response_class=HTMLResponse)
async def confirm_delete_group_subject(group_id: int, subject_id: int, db: AsyncSession = Depends(get_db)):
    html = f"""
    <h1>Удалить связь группы {group_id} и предмета {subject_id}?</h1>
    <form method="post" action="/admin/group-subjects/delete/{group_id}/{subject_id}">
        <button type="submit">Да, удалить</button>
        <a href="/admin/group-subjects">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/group-subjects/delete/{group_id}/{subject_id}", response_model=None)
async def delete_group_subject(group_id: int, subject_id: int, db: AsyncSession = Depends(get_db)):
    await group_subject_service.delete_group_subject(db, group_id, subject_id)
    return RedirectResponse(url="/admin/group-subjects", status_code=303)

# ---------- ЭКСПОРТ В EXCEL (все маршруты) ----------
@router.get("/export/groups")
async def export_groups(db: AsyncSession = Depends(get_db)):
    groups = await group_service.get_all_groups(db)
    faculties = await faculty_service.get_all_faculties(db)
    fac_dict = {f.id: f.name for f in faculties}
    data = [{"ID": g.id, "Название": g.name, "Факультет ID": g.faculty_id, "Факультет": fac_dict.get(g.faculty_id, ""), "Студентов": g.student_count} for g in groups]
    return export_to_excel(data, ["ID", "Название", "Факультет ID", "Факультет", "Студентов"], "Группы", "groups.xlsx")

@router.get("/export/teachers")
async def export_teachers(db: AsyncSession = Depends(get_db)):
    teachers = await teacher_service.get_all_teachers(db)
    data = [{"ID": t.id, "Логин": t.login, "ФИО": t.name, "Ссылка": t.url or ""} for t in teachers]
    return export_to_excel(data, ["ID", "Логин", "ФИО", "Ссылка"], "Преподаватели", "teachers.xlsx")

@router.get("/export/audiences")
async def export_audiences(db: AsyncSession = Depends(get_db)):
    audiences = await audience_service.get_all_audiences(db)
    buildings = await building_service.get_all_buildings(db)
    b_dict = {b.id: b.name for b in buildings}
    data = [{"ID": a.id, "Название": a.name, "Здание ID": a.building_id, "Здание": b_dict.get(a.building_id, ""), "Вместимость": a.capacity, "Тип": a.type or "", "Активна": "Да" if a.is_active else "Нет"} for a in audiences]
    return export_to_excel(data, ["ID", "Название", "Здание ID", "Здание", "Вместимость", "Тип", "Активна"], "Аудитории", "audiences.xlsx")

@router.get("/export/buildings")
async def export_buildings(db: AsyncSession = Depends(get_db)):
    buildings = await building_service.get_all_buildings(db)
    data = [{"ID": b.id, "Название": b.name, "Адрес": b.address or ""} for b in buildings]
    return export_to_excel(data, ["ID", "Название", "Адрес"], "Здания", "buildings.xlsx")

@router.get("/export/calendar")
async def export_calendar(db: AsyncSession = Depends(get_db)):
    entries = await calendar_service.get_all_calendar(db)
    data = [{"ID": e.id, "Дата": e.date, "Рабочий день": "Да" if e.is_working_day else "Нет", "Тип недели": e.week_type or "", "Описание": e.description or ""} for e in entries]
    return export_to_excel(data, ["ID", "Дата", "Рабочий день", "Тип недели", "Описание"], "Календарь", "calendar.xlsx")

@router.get("/export/faculties")
async def export_faculties(db: AsyncSession = Depends(get_db)):
    faculties = await faculty_service.get_all_faculties(db)
    data = [{"ID": f.id, "Название": f.name, "Отображаемое имя": f.display_name or "", "Короткое имя": f.short_display_name or ""} for f in faculties]
    return export_to_excel(data, ["ID", "Название", "Отображаемое имя", "Короткое имя"], "Факультеты", "faculties.xlsx")

@router.get("/export/user_types")
async def export_user_types(db: AsyncSession = Depends(get_db)):
    types = await user_type_service.get_all_user_types(db)
    data = [{"ID": t.id, "Код": t.code, "Название": t.name, "Описание": t.description or ""} for t in types]
    return export_to_excel(data, ["ID", "Код", "Название", "Описание"], "Типы пользователей", "user_types.xlsx")

@router.get("/export/subdivisions")
async def export_subdivisions(db: AsyncSession = Depends(get_db)):
    subs = await subdivision_service.get_all_subdivisions(db)
    data = [{"ID": s.id, "Название": s.name} for s in subs]
    return export_to_excel(data, ["ID", "Название"], "Подразделения", "subdivisions.xlsx")

@router.get("/export/roles")
async def export_roles(db: AsyncSession = Depends(get_db)):
    roles = await role_service.get_all_roles(db)
    data = [{"ID": r.id, "Название": r.role_name, "Описание": r.description or ""} for r in roles]
    return export_to_excel(data, ["ID", "Название", "Описание"], "Роли", "roles.xlsx")

@router.get("/export/permissions")
async def export_permissions(db: AsyncSession = Depends(get_db)):
    perms = await permission_service.get_all_permissions(db)
    data = [{"ID": p.id, "Код": p.permission_code, "Название": p.permission_name, "Описание": p.description or ""} for p in perms]
    return export_to_excel(data, ["ID", "Код", "Название", "Описание"], "Права", "permissions.xlsx")

@router.get("/export/time_slots")
async def export_time_slots(db: AsyncSession = Depends(get_db)):
    slots = await time_slot_service.get_all_time_slots(db)
    data = [{"ID": s.id, "Номер": s.slot_number, "Название": s.name, "Начало": str(s.start_time), "Конец": str(s.end_time), "Длительность(мин)": s.duration_minutes or "", "Перерыв(мин)": s.break_after_minutes or "", "Активен": "Да" if s.is_active else "Нет"} for s in slots]
    return export_to_excel(data, ["ID", "Номер", "Название", "Начало", "Конец", "Длительность(мин)", "Перерыв(мин)", "Активен"], "Временные слоты", "time_slots.xlsx")

@router.get("/export/subjects")
async def export_subjects(db: AsyncSession = Depends(get_db)):
    subjects = await subject_service.get_all_subjects(db)
    data = [{"ID": s.id, "Название": s.name} for s in subjects]
    return export_to_excel(data, ["ID", "Название"], "Дисциплины", "subjects.xlsx")

@router.get("/export/teacher_groups")
async def export_teacher_groups(db: AsyncSession = Depends(get_db)):
    rels = await teacher_group_service.get_all_teacher_groups(db)
    teachers = await teacher_service.get_all_teachers(db)
    groups = await group_service.get_all_groups(db)
    t_dict = {t.id: t.name for t in teachers}
    g_dict = {g.id: g.name for g in groups}
    data = [{"ID учителя": r['teacher_id'], "Учитель": t_dict.get(r['teacher_id'], ""), "ID группы": r['group_id'], "Группа": g_dict.get(r['group_id'], "")} for r in rels]
    return export_to_excel(data, ["ID учителя", "Учитель", "ID группы", "Группа"], "Преподаватели-группы", "teacher_groups.xlsx")

@router.get("/export/teacher_subjects")
async def export_teacher_subjects(db: AsyncSession = Depends(get_db)):
    rels = await teacher_subject_service.get_all_teacher_subjects(db)
    teachers = await teacher_service.get_all_teachers(db)
    subjects = await subject_service.get_all_subjects(db)
    t_dict = {t.id: t.name for t in teachers}
    s_dict = {s.id: s.name for s in subjects}
    data = [{"ID учителя": r['teacher_id'], "Учитель": t_dict.get(r['teacher_id'], ""), "ID предмета": r['subject_id'], "Предмет": s_dict.get(r['subject_id'], ""), "Основной": "Да" if r['is_main'] else "Нет"} for r in rels]
    return export_to_excel(data, ["ID учителя", "Учитель", "ID предмета", "Предмет", "Основной"], "Преподаватели-предметы", "teacher_subjects.xlsx")

@router.get("/export/audience_subjects")
async def export_audience_subjects(db: AsyncSession = Depends(get_db)):
    rels = await audience_subject_service.get_all_audience_subjects(db)
    audiences = await audience_service.get_all_audiences(db)
    subjects = await subject_service.get_all_subjects(db)
    a_dict = {a.id: a.name for a in audiences}
    s_dict = {s.id: s.name for s in subjects}
    data = [{"ID аудитории": r['audience_id'], "Аудитория": a_dict.get(r['audience_id'], ""), "ID предмета": r['subject_id'], "Предмет": s_dict.get(r['subject_id'], "")} for r in rels]
    return export_to_excel(data, ["ID аудитории", "Аудитория", "ID предмета", "Предмет"], "Аудитории-предметы", "audience_subjects.xlsx")

@router.get("/export/group_subjects")
async def export_group_subjects(db: AsyncSession = Depends(get_db)):
    rels = await group_subject_service.get_all_group_subjects(db)
    groups = await group_service.get_all_groups(db)
    subjects = await subject_service.get_all_subjects(db)
    g_dict = {g.id: g.name for g in groups}
    s_dict = {s.id: s.name for s in subjects}
    data = [{"ID группы": r['group_id'], "Группа": g_dict.get(r['group_id'], ""), "ID предмета": r['subject_id'], "Предмет": s_dict.get(r['subject_id'], "")} for r in rels]
    return export_to_excel(data, ["ID группы", "Группа", "ID предмета", "Предмет"], "Группы-предметы", "group_subjects.xlsx")

# ---------- РУЧНОЕ СОСТАВЛЕНИЕ РАСПИСАНИЯ ----------
from datetime import date
from sqlalchemy import delete
from app.models.reference import Lesson
import io
from fastapi import UploadFile, File
from openpyxl import load_workbook
from datetime import datetime

@router.get("/manual-schedule", response_class=HTMLResponse)
async def manual_schedule_page(db: AsyncSession = Depends(get_db)):
    subjects = await subject_service.get_all_subjects(db)
    time_slots = await time_slot_service.get_all_time_slots(db)
    lessons = await db.execute(select(Lesson).order_by(Lesson.date))
    lessons = lessons.scalars().all()

    lesson_data = []
    for les in lessons:
        lesson_data.append({
    "id": les.id,
    "date": les.date,
    "week_day": les.week_day,
    "subject": les.subject.name if les.subject else "—",
    "time_slot": les.time_slot.name if les.time_slot else "—",
    "student_count": les.student_count or ""
}) 

    html = f"""
    <html>
    <head>
        <title>Ручное составление расписания</title>
        <style>
            body {{ font-family: sans-serif; margin: 20px; }}
            .form-group {{ margin-bottom: 10px; }}
            label {{ display: inline-block; width: 120px; }}
            .btn {{ background: #007bff; color: white; padding: 6px 12px; border: none; border-radius: 4px; cursor: pointer; }}
            .btn-danger {{ background: #dc3545; }}
            .btn-excel {{ background: #28a745; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background: #f2f2f2; }}
            .delete-form {{ display: inline; }}
        </style>
    </head>
    <body>
        <h1>Ручное составление расписания</h1>
        <form method="post" action="/admin/manual-schedule/add">
            <div class="form-group">
                <label>Предмет:</label>
                <select name="subject_id" required>
                    <option value="">Выберите предмет</option>
    """
    for s in subjects:
        html += f'<option value="{s.id}">{s.name}</option>'
    html += """
                </select>
            </div>
            <div class="form-group">
                <label>Дата:</label>
                <input type="date" name="date" required>
            </div>
            <div class="form-group">
                <label>Время (пара):</label>
                <select name="time_slot_id" required>
                    <option value="">Выберите время</option>
    """
    for ts in time_slots:
        html += f'<option value="{ts.id}">{ts.name} ({ts.start_time}-{ts.end_time})</option>'
    html += """
                </select>
            </div>
            <div class="form-group">
                <label>Кол-во студентов:</label>
                <input type="number" name="student_count" placeholder="Необязательно">
            </div>
            <button type="submit" class="btn">Добавить занятие</button>
        </form>

        <hr>
        <h2>Загрузить из Excel</h2>
        <form method="post" action="/admin/manual-schedule/upload" enctype="multipart/form-data">
            <input type="file" name="file" accept=".xlsx,.xls" required>
            <button type="submit" class="btn btn-excel">📎 Загрузить Excel</button>
        </form>
        <p style="font-size: 0.9em; color: #666;">Формат: первая строка — заголовки: subject_name, date, time_slot_name, student_count, type, text. 
        Предметы и временные слоты должны существовать в справочниках.</p>
        <hr>

        <h2>Созданные занятия</h2>
        <form method="post" action="/admin/manual-schedule/bulk-delete" onsubmit="return confirmDeleteSelected();">
            <button type="submit" class="btn btn-danger">Удалить выбранные</button>
            <label><input type="checkbox" id="selectAllLessons"> Выделить всё</label>
            <table>
                <thead>
                    <tr><th>Выбрать</th><th>ID</th><th>Дата</th><th>Предмет</th><th>Время</th><th>Студентов</th><th>Действия</th></tr>
                </thead>
                <tbody>
    """
    for ld in lesson_data:
        html += f"""
            <tr>
                <td><input type="checkbox" name="ids" value="{ld['id']}"></td>
                <td>{ld['id']}</td>
                <td>{ld['date']}</td>
                <td>{ld['week_day']}</td>
                <td>{ld['subject']}</td>
                <td>{ld['time_slot']}</td>
                <td>{ld['student_count']}</td>
                <td>
                    <a href="/admin/manual-schedule/delete/{ld['id']}" class="btn btn-danger" style="text-decoration:none; padding:2px 8px;">Удалить</a>
                </td>
            </tr>
        """
    html += """
                </tbody>
            </table>
        </form>
        <br><a href="/admin/">На главную админки</a>
        <script>
            const selectAll = document.getElementById('selectAllLessons');
            if(selectAll) selectAll.addEventListener('change', function() {
                document.querySelectorAll('input[name="ids"]').forEach(cb => cb.checked = selectAll.checked);
            });
            function confirmDeleteSelected() {
                const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
                if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
                return confirm('Удалить выбранные занятия?');
            }
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

@router.post("/manual-schedule/add", response_model=None)
async def add_manual_lesson(
    subject_id: int = Form(...),
    date: date = Form(...),
    time_slot_id: int = Form(...),
    student_count: int = Form(None),
    db: AsyncSession = Depends(get_db)
):
    week_day = date.isoweekday()   # вычисляем день недели
    lesson = Lesson(
        subject_id=subject_id,
        date=date,
        time_slot_id=time_slot_id,
        student_count=student_count,
        week_day=week_day
    )
    db.add(lesson)
    await db.commit()
    return RedirectResponse(url="/admin/manual-schedule", status_code=303)

@router.post("/manual-schedule/bulk-delete", response_model=None)
async def bulk_delete_lessons(ids: List[int] = Form(...), db: AsyncSession = Depends(get_db)):
    for lid in ids:
        await db.execute(delete(Lesson).where(Lesson.id == lid))
    await db.commit()
    return RedirectResponse(url="/admin/manual-schedule", status_code=303)

@router.get("/manual-schedule/delete/{lesson_id}", response_class=HTMLResponse)
async def confirm_delete_lesson(lesson_id: int, db: AsyncSession = Depends(get_db)):
    html = f"""
    <h1>Удалить занятие {lesson_id}?</h1>
    <form method="post" action="/admin/manual-schedule/delete/{lesson_id}">
        <button type="submit" class="btn btn-danger">Да, удалить</button>
        <a href="/admin/manual-schedule">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/manual-schedule/delete/{lesson_id}", response_model=None)
async def delete_lesson(lesson_id: int, db: AsyncSession = Depends(get_db)):
    await db.execute(delete(Lesson).where(Lesson.id == lesson_id))
    await db.commit()
    return RedirectResponse(url="/admin/manual-schedule", status_code=303)

@router.post("/manual-schedule/upload")
async def upload_lessons_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Файл должен быть Excel (.xlsx или .xls)")

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Ошибка чтения файла: {e}")

    expected_headers = ['subject_name', 'date', 'time_slot_name', 'student_count', 'type', 'text']
    headers = [cell.value for cell in ws[1]]
    missing = set(expected_headers) - set(headers)
    if missing:
        raise HTTPException(400, f"Отсутствуют колонки: {missing}")

    col_idx = {h: headers.index(h) for h in expected_headers}
    subjects = {s.name: s.id for s in await subject_service.get_all_subjects(db)}
    time_slots = {ts.name: ts.id for ts in await time_slot_service.get_all_time_slots(db)}

    added = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            subject_name = row[col_idx['subject_name']]
            date_str = row[col_idx['date']]
            time_slot_name = row[col_idx['time_slot_name']]
            student_count = row[col_idx['student_count']]
            lesson_type = row[col_idx['type']] if col_idx['type'] < len(row) else None
            text = row[col_idx['text']] if col_idx['text'] < len(row) else None

            if isinstance(date_str, datetime):
                lesson_date = date_str.date()
            else:
                lesson_date = datetime.strptime(str(date_str), "%Y-%m-%d").date()

            if subject_name not in subjects:
                errors.append(f"Строка {row_idx}: предмет '{subject_name}' не найден")
                continue
            subject_id = subjects[subject_name]

            if time_slot_name not in time_slots:
                errors.append(f"Строка {row_idx}: временной слот '{time_slot_name}' не найден")
                continue
            time_slot_id = time_slots[time_slot_name]

            week_day = lesson_date.isoweekday()

            lesson = Lesson(
                subject_id=subject_id,
                date=lesson_date,
                time_slot_id=time_slot_id,
                student_count=student_count if student_count else None,
                type=lesson_type,
                text=text,
                week_day=week_day
            )
            db.add(lesson)
            added += 1
        except Exception as e:
            errors.append(f"Строка {row_idx}: {str(e)}")

    await db.commit()
    return {
        "message": "Загрузка завершена",
        "added": added,
        "errors": errors
    }

# ---------- СТАТИСТИКА ----------
@router.get("/stats", response_class=HTMLResponse)
async def stats_page(db: AsyncSession = Depends(get_db)):
    groups_count = len(await group_service.get_all_groups(db))
    teachers_count = len(await teacher_service.get_all_teachers(db))
    audiences_count = len(await audience_service.get_all_audiences(db))
    buildings_count = len(await building_service.get_all_buildings(db))
    calendar_count = len(await calendar_service.get_all_calendar(db))
    faculties_count = len(await faculty_service.get_all_faculties(db))
    user_types_count = len(await user_type_service.get_all_user_types(db))
    subdivisions_count = len(await subdivision_service.get_all_subdivisions(db))
    roles_count = len(await role_service.get_all_roles(db))
    permissions_count = len(await permission_service.get_all_permissions(db))
    time_slots_count = len(await time_slot_service.get_all_time_slots(db))
    subjects_count = len(await subject_service.get_all_subjects(db))
    teacher_groups_count = len(await teacher_group_service.get_all_teacher_groups(db))
    teacher_subjects_count = len(await teacher_subject_service.get_all_teacher_subjects(db))
    audience_subjects_count = len(await audience_subject_service.get_all_audience_subjects(db))
    group_subjects_count = len(await group_subject_service.get_all_group_subjects(db))

    html = f"""
    <html>
    <head><title>Статистика системы</title>
    <style>
        body {{ font-family: sans-serif; margin: 20px; }}
        h1 {{ color: #2c3e50; }}
        .stats-container {{ display: flex; flex-wrap: wrap; gap: 20px; }}
        .card {{ background: #f8f9fa; border-radius: 8px; padding: 15px; width: 250px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .card h3 {{ margin: 0 0 10px; color: #3498db; }}
        .card .count {{ font-size: 32px; font-weight: bold; color: #2c3e50; }}
        hr {{ margin: 20px 0; }}
    </style>
    </head>
    <body>
        <h1>Статистика системы</h1>
        <div class='stats-container'>
            <div class='card'><h3>Группы</h3><div class='count'>{groups_count}</div></div>
            <div class='card'><h3>Преподаватели</h3><div class='count'>{teachers_count}</div></div>
            <div class='card'><h3>Аудитории</h3><div class='count'>{audiences_count}</div></div>
            <div class='card'><h3>Здания</h3><div class='count'>{buildings_count}</div></div>
            <div class='card'><h3>Календарь</h3><div class='count'>{calendar_count}</div></div>
            <div class='card'><h3>Факультеты</h3><div class='count'>{faculties_count}</div></div>
            <div class='card'><h3>Типы пользователей</h3><div class='count'>{user_types_count}</div></div>
            <div class='card'><h3>Подразделения</h3><div class='count'>{subdivisions_count}</div></div>
            <div class='card'><h3>Роли</h3><div class='count'>{roles_count}</div></div>
            <div class='card'><h3>Права</h3><div class='count'>{permissions_count}</div></div>
            <div class='card'><h3>Временные слоты</h3><div class='count'>{time_slots_count}</div></div>
            <div class='card'><h3>Дисциплины</h3><div class='count'>{subjects_count}</div></div>
        </div>
        <hr>
        <h2>Связи</h2>
        <div class='stats-container'>
            <div class='card'><h3>Преподаватели ↔ Группы</h3><div class='count'>{teacher_groups_count}</div></div>
            <div class='card'><h3>Преподаватели ↔ Предметы</h3><div class='count'>{teacher_subjects_count}</div></div>
            <div class='card'><h3>Аудитории ↔ Предметы</h3><div class='count'>{audience_subjects_count}</div></div>
            <div class='card'><h3>Группы ↔ Предметы</h3><div class='count'>{group_subjects_count}</div></div>
        </div>
        <br><a href='/admin/'>На главную админки</a>
    </body>
    </html>
    """
    return HTMLResponse(content=html)