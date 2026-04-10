import io
import json
from typing import List
from fastapi import Request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse

from app.services.audit import log_action

def export_to_excel(data: list, columns: list, sheet_name: str, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(columns, 1):
            value = row_data.get(key)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def _common_styles():
    return """
    <style>
        .btn-excel {
            display: inline-block;
            background: #28a745;
            color: white;
            padding: 6px 12px;
            text-decoration: none;
            border-radius: 4px;
            margin-bottom: 15px;
        }
        .btn-import {
            background: #007bff;
            margin-left: 10px;
        }
        table { border-collapse: collapse; width: 100%; margin-top: 10px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        .import-form { display: none; margin: 10px 0; padding: 10px; border: 1px solid #ccc; background: #f9f9f9; }
    </style>
    """

def _import_script(entity: str):
    return f"""
    <script>
        function toggleImportForm_{entity}() {{
            var form = document.getElementById('importForm_{entity}');
            if (form.style.display === 'none' || form.style.display === '') {{
                form.style.display = 'block';
            }} else {{
                form.style.display = 'none';
            }}
        }}
        document.getElementById('uploadForm_{entity}').onsubmit = async function(e) {{
            e.preventDefault();
            const formData = new FormData(this);
            const response = await fetch('/admin/import/{entity}', {{ method: 'POST', body: formData }});
            const result = await response.json();
            alert(`Добавлено: ${{result.added}}, Обновлено: ${{result.updated}}\\nОшибки: ${{result.errors.length}}`);
            window.location.reload();
        }};
    </script>
    """

# app/routers/admin/common.py
import io
import json
from typing import List
from fastapi import Request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.services.audit import log_action

def export_to_excel(data: list, columns: list, sheet_name: str, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(columns, 1):
            value = row_data.get(key)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def _common_styles():
    return """
    <style>
        .btn-excel {
            display: inline-block;
            background: #28a745;
            color: white;
            padding: 6px 12px;
            text-decoration: none;
            border-radius: 4px;
            margin-bottom: 15px;
        }
        .btn-import {
            background: #007bff;
            margin-left: 10px;
        }
        table { border-collapse: collapse; width: 100%; margin-top: 10px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        .import-form { display: none; margin: 10px 0; padding: 10px; border: 1px solid #ccc; background: #f9f9f9; }
    </style>
    """

def _import_script(entity: str):
    return f"""
    <script>
        function toggleImportForm_{entity}() {{
            var form = document.getElementById('importForm_{entity}');
            if (form.style.display === 'none' || form.style.display === '') {{
                form.style.display = 'block';
            }} else {{
                form.style.display = 'none';
            }}
        }}
        document.getElementById('uploadForm_{entity}').onsubmit = async function(e) {{
            e.preventDefault();
            const formData = new FormData(this);
            const response = await fetch('/admin/import/{entity}', {{ method: 'POST', body: formData }});
            const result = await response.json();
            alert(`Добавлено: ${{result.added}}, Обновлено: ${{result.updated}}\\nОшибки: ${{result.errors.length}}`);
            window.location.reload();
        }};
    </script>
    """

# Вспомогательные функции для проверки зависимостей (общие для разных сущностей)
async def has_dependent_groups(db: AsyncSession, faculty_id: int) -> bool:
    from app.models.reference import Group
    result = await db.execute(select(func.count()).select_from(Group).where(Group.faculty_id == faculty_id))
    return result.scalar() > 0

async def has_dependent_audiences(db: AsyncSession, building_id: int) -> bool:
    from app.models.reference import Audience
    result = await db.execute(select(func.count()).select_from(Audience).where(Audience.building_id == building_id))
    return result.scalar() > 0

async def has_dependent_audience_subjects(db: AsyncSession, audience_id: int) -> bool:
    from app.models.reference import op_audiences_of_pairs
    result = await db.execute(select(func.count()).select_from(op_audiences_of_pairs).where(op_audiences_of_pairs.c.audience_id == audience_id))
    return result.scalar() > 0

async def has_dependent_teacher_groups(db: AsyncSession, teacher_id: int) -> bool:
    from app.models.reference import op_teachers_groups
    result = await db.execute(select(func.count()).select_from(op_teachers_groups).where(op_teachers_groups.c.teachers_id == teacher_id))
    return result.scalar() > 0

async def has_dependent_teacher_subjects(db: AsyncSession, teacher_id: int) -> bool:
    from app.models.reference import op_teachers_of_pairs
    result = await db.execute(select(func.count()).select_from(op_teachers_of_pairs).where(op_teachers_of_pairs.c.teacher_id == teacher_id))
    return result.scalar() > 0

async def has_dependent_group_teachers(db: AsyncSession, group_id: int) -> bool:
    from app.models.reference import op_teachers_groups
    result = await db.execute(select(func.count()).select_from(op_teachers_groups).where(op_teachers_groups.c.groups_id == group_id))
    return result.scalar() > 0

async def has_dependent_group_subjects(db: AsyncSession, group_id: int) -> bool:
    from app.models.reference import op_groups_of_pairs
    result = await db.execute(select(func.count()).select_from(op_groups_of_pairs).where(op_groups_of_pairs.c.group_id == group_id))
    return result.scalar() > 0

async def has_dependent_subject_teachers(db: AsyncSession, subject_id: int) -> bool:
    from app.models.reference import op_teachers_of_pairs
    result = await db.execute(select(func.count()).select_from(op_teachers_of_pairs).where(op_teachers_of_pairs.c.subject_id == subject_id))
    return result.scalar() > 0

async def has_dependent_subject_audiences(db: AsyncSession, subject_id: int) -> bool:
    from app.models.reference import op_audiences_of_pairs
    result = await db.execute(select(func.count()).select_from(op_audiences_of_pairs).where(op_audiences_of_pairs.c.subject_id == subject_id))
    return result.scalar() > 0

async def has_dependent_subject_groups(db: AsyncSession, subject_id: int) -> bool:
    from app.models.reference import op_groups_of_pairs
    result = await db.execute(select(func.count()).select_from(op_groups_of_pairs).where(op_groups_of_pairs.c.subject_id == subject_id))
    return result.scalar() > 0

def _common_styles():
    return """
    <style>
        .btn-excel {
            display: inline-block;
            background: #28a745;
            color: white;
            padding: 6px 12px;
            text-decoration: none;
            border-radius: 4px;
            margin-bottom: 15px;
        }
        .btn-import {
            background: #007bff;
            margin-left: 10px;
        }
        table { border-collapse: collapse; width: 100%; margin-top: 10px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        .import-form { display: none; margin: 10px 0; padding: 10px; border: 1px solid #ccc; background: #f9f9f9; }
    </style>
    """

# Проверки зависимостей для факультетов, зданий, дисциплин и т.д.
async def has_dependent_faculty_groups(db: AsyncSession, faculty_id: int) -> bool:
    from app.models.reference import Group
    result = await db.execute(select(func.count()).select_from(Group).where(Group.faculty_id == faculty_id))
    return result.scalar() > 0

async def has_dependent_building_audiences(db: AsyncSession, building_id: int) -> bool:
    from app.models.reference import Audience
    result = await db.execute(select(func.count()).select_from(Audience).where(Audience.building_id == building_id))
    return result.scalar() > 0

async def has_dependent_subject_teachers(db: AsyncSession, subject_id: int) -> bool:
    from app.models.reference import op_teachers_of_pairs
    result = await db.execute(select(func.count()).select_from(op_teachers_of_pairs).where(op_teachers_of_pairs.c.subject_id == subject_id))
    return result.scalar() > 0

async def has_dependent_subject_audiences(db: AsyncSession, subject_id: int) -> bool:
    from app.models.reference import op_audiences_of_pairs
    result = await db.execute(select(func.count()).select_from(op_audiences_of_pairs).where(op_audiences_of_pairs.c.subject_id == subject_id))
    return result.scalar() > 0

async def has_dependent_subject_groups(db: AsyncSession, subject_id: int) -> bool:
    from app.models.reference import op_groups_of_pairs
    result = await db.execute(select(func.count()).select_from(op_groups_of_pairs).where(op_groups_of_pairs.c.subject_id == subject_id))
    return result.scalar() > 0

async def has_dependent_teacher_groups(db: AsyncSession, teacher_id: int) -> bool:
    from app.models.reference import op_teachers_groups
    result = await db.execute(select(func.count()).select_from(op_teachers_groups).where(op_teachers_groups.c.teachers_id == teacher_id))
    return result.scalar() > 0

async def has_dependent_teacher_subjects(db: AsyncSession, teacher_id: int) -> bool:
    from app.models.reference import op_teachers_of_pairs
    result = await db.execute(select(func.count()).select_from(op_teachers_of_pairs).where(op_teachers_of_pairs.c.teacher_id == teacher_id))
    return result.scalar() > 0

async def has_dependent_audience_subjects(db: AsyncSession, audience_id: int) -> bool:
    from app.models.reference import op_audiences_of_pairs
    result = await db.execute(select(func.count()).select_from(op_audiences_of_pairs).where(op_audiences_of_pairs.c.audience_id == audience_id))
    return result.scalar() > 0

async def has_dependent_group_subjects(db: AsyncSession, group_id: int) -> bool:
    from app.models.reference import op_groups_of_pairs
    result = await db.execute(select(func.count()).select_from(op_groups_of_pairs).where(op_groups_of_pairs.c.group_id == group_id))
    return result.scalar() > 0

async def has_dependent_group_teachers(db: AsyncSession, group_id: int) -> bool:
    from app.models.reference import op_teachers_groups
    result = await db.execute(select(func.count()).select_from(op_teachers_groups).where(op_teachers_groups.c.groups_id == group_id))
    return result.scalar() > 0