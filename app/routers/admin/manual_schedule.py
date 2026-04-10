from datetime import date, datetime
from typing import List
from fastapi import APIRouter, Depends, Form, Request, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from app.core.database import get_db
from app.services import subject as subject_service
from app.services import time_slot as time_slot_service
from app.models.reference import Lesson
from app.routers.admin.common import _common_styles
from app.services.audit import log_action
import io
from openpyxl import load_workbook

router = APIRouter(prefix="/admin/manual-schedule", tags=["Admin Manual Schedule"])

@router.get("", response_class=HTMLResponse)
async def manual_schedule_page(db: AsyncSession = Depends(get_db)):
    subjects = await subject_service.get_all_subjects(db)
    time_slots = await time_slot_service.get_all_time_slots(db)
    lessons = await db.execute(select(Lesson).order_by(Lesson.date))
    lessons = lessons.scalars().all()

    lesson_data = []
    for les in lessons:
        lesson_data.append({
            "id": les.id,
            "date": les.date,
            "subject": les.subject.name if les.subject else "—",
            "time_slot": les.time_slot.name if les.time_slot else "—",
            "student_count": les.student_count or ""
        })

    html = f"""
    <html>
    <head>
        <title>Ручное составление расписания</title>
        {_common_styles()}
        <style>
            .form-group {{ margin-bottom: 10px; }}
            label {{ display: inline-block; width: 120px; }}
            .btn {{ background: #007bff; color: white; padding: 6px 12px; border: none; border-radius: 4px; cursor: pointer; }}
            .btn-danger {{ background: #dc3545; }}
            .btn-excel {{ background: #28a745; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background: #f2f2f2; }}
            .delete-form {{ display: inline; }}
        </style>
    </head>
    <body>
        <h1>Ручное составление расписания</h1>
        <form method="post" action="/admin/manual-schedule/add">
            <div class="form-group">
                <label>Предмет:</label>
                <select name="subject_id" required>
                    <option value="">Выберите предмет</option>
    """
    for s in subjects:
        html += f'<option value="{s.id}">{s.name}</option>'
    html += """
                </select>
            </div>
            <div class="form-group">
                <label>Дата:</label>
                <input type="date" name="date" required>
            </div>
            <div class="form-group">
                <label>Время (пара):</label>
                <select name="time_slot_id" required>
                    <option value="">Выберите время</option>
    """
    for ts in time_slots:
        html += f'<option value="{ts.id}">{ts.name} ({ts.start_time}-{ts.end_time})</option>'
    html += """
                </select>
            </div>
            <div class="form-group">
                <label>Кол-во студентов:</label>
                <input type="number" name="student_count" placeholder="Необязательно">
            </div>
            <button type="submit" class="btn">Добавить занятие</button>
        </form>

        <hr>
        <h2>Загрузить из Excel</h2>
        <form method="post" action="/admin/manual-schedule/upload" enctype="multipart/form-data">
            <input type="file" name="file" accept=".xlsx,.xls" required>
            <button type="submit" class="btn btn-excel">📎 Загрузить Excel</button>
        </form>
        <p style="font-size: 0.9em; color: #666;">Формат: первая строка — заголовки: subject_name, date, time_slot_name, student_count, type, text. 
        Предметы и временные слоты должны существовать в справочниках.</p>
        <hr>

        <h2>Созданные занятия</h2>
        <form method="post" action="/admin/manual-schedule/bulk-delete" onsubmit="return confirmDeleteSelected();">
            <button type="submit" class="btn btn-danger">Удалить выбранные</button>
            <label><input type="checkbox" id="selectAllLessons"> Выделить всё</label>
            <table>
                <thead>
                    <tr><th>Выбрать</th><th>ID</th><th>Дата</th><th>Предмет</th><th>Время</th><th>Студентов</th><th>Действия</th></tr>
                </thead>
                <tbody>
    """
    for ld in lesson_data:
        html += f"""
            <tr>
                <td><input type="checkbox" name="ids" value="{ld['id']}"></td>
                <td>{ld['id']}</td>
                <td>{ld['date']}</td>
                <td>{ld['subject']}</td>
                <td>{ld['time_slot']}</td>
                <td>{ld['student_count']}</td>
                <td>
                    <a href="/admin/manual-schedule/delete/{ld['id']}" class="btn btn-danger" style="text-decoration:none; padding:2px 8px;">Удалить</a>
                </td>
            </tr>
        """
    html += """
                </tbody>
            </table>
        </form>
        <br><a href="/admin/">На главную админки</a>
        <script>
            const selectAll = document.getElementById('selectAllLessons');
            if(selectAll) selectAll.addEventListener('change', function() {
                document.querySelectorAll('input[name="ids"]').forEach(cb => cb.checked = selectAll.checked);
            });
            function confirmDeleteSelected() {
                const anyChecked = document.querySelectorAll('input[name="ids"]:checked').length > 0;
                if (!anyChecked) { alert('Не выбрано ни одной записи'); return false; }
                return confirm('Удалить выбранные занятия?');
            }
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

@router.post("/add", response_model=None)
async def add_manual_lesson(
    request: Request,
    subject_id: int = Form(...),
    date: date = Form(...),
    time_slot_id: int = Form(...),
    student_count: int = Form(None),
    db: AsyncSession = Depends(get_db)
):
    week_day = date.isoweekday()
    lesson = Lesson(
        subject_id=subject_id,
        date=date,
        time_slot_id=time_slot_id,
        student_count=student_count,
        week_day=week_day
    )
    db.add(lesson)
    await db.commit()
    await log_action(db, "create_manual_lesson", {"subject_id": subject_id, "date": str(date), "time_slot_id": time_slot_id, "student_count": student_count}, request)
    return RedirectResponse(url="/admin/manual-schedule", status_code=303)

@router.post("/bulk-delete", response_model=None)
async def bulk_delete_lessons(
    request: Request,
    ids: List[int] = Form(...),
    db: AsyncSession = Depends(get_db)
):
    for lid in ids:
        await db.execute(delete(Lesson).where(Lesson.id == lid))
    await db.commit()
    await log_action(db, "bulk_delete_manual_lessons", {"ids": ids}, request)
    return RedirectResponse(url="/admin/manual-schedule", status_code=303)

@router.get("/delete/{lesson_id}", response_class=HTMLResponse)
async def confirm_delete_lesson(lesson_id: int, db: AsyncSession = Depends(get_db)):
    html = f"""
    <h1>Удалить занятие {lesson_id}?</h1>
    <form method="post" action="/admin/manual-schedule/delete/{lesson_id}">
        <button type="submit" class="btn btn-danger">Да, удалить</button>
        <a href="/admin/manual-schedule">Отмена</a>
    </form>
    """
    return HTMLResponse(content=html)

@router.post("/delete/{lesson_id}", response_model=None)
async def delete_lesson(
    request: Request,
    lesson_id: int,
    db: AsyncSession = Depends(get_db)
):
    await db.execute(delete(Lesson).where(Lesson.id == lesson_id))
    await db.commit()
    await log_action(db, "delete_manual_lesson", {"id": lesson_id}, request)
    return RedirectResponse(url="/admin/manual-schedule", status_code=303)

@router.post("/upload")
async def upload_lessons_excel(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Файл должен быть Excel (.xlsx или .xls)")

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Ошибка чтения файла: {e}")

    expected_headers = ['subject_name', 'date', 'time_slot_name', 'student_count', 'type', 'text']
    headers = [cell.value for cell in ws[1]]
    missing = set(expected_headers) - set(headers)
    if missing:
        raise HTTPException(400, f"Отсутствуют колонки: {missing}")

    col_idx = {h: headers.index(h) for h in expected_headers}
    subjects = {s.name: s.id for s in await subject_service.get_all_subjects(db)}
    time_slots = {ts.name: ts.id for ts in await time_slot_service.get_all_time_slots(db)}

    added = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            subject_name = row[col_idx['subject_name']]
            date_str = row[col_idx['date']]
            time_slot_name = row[col_idx['time_slot_name']]
            student_count = row[col_idx['student_count']]
            lesson_type = row[col_idx['type']] if col_idx['type'] < len(row) else None
            text = row[col_idx['text']] if col_idx['text'] < len(row) else None

            if isinstance(date_str, datetime):
                lesson_date = date_str.date()
            else:
                lesson_date = datetime.strptime(str(date_str), "%Y-%m-%d").date()

            if subject_name not in subjects:
                errors.append(f"Строка {row_idx}: предмет '{subject_name}' не найден")
                continue
            subject_id = subjects[subject_name]

            if time_slot_name not in time_slots:
                errors.append(f"Строка {row_idx}: временной слот '{time_slot_name}' не найден")
                continue
            time_slot_id = time_slots[time_slot_name]

            week_day = lesson_date.isoweekday()

            lesson = Lesson(
                subject_id=subject_id,
                date=lesson_date,
                time_slot_id=time_slot_id,
                student_count=student_count if student_count else None,
                type=lesson_type,
                text=text,
                week_day=week_day
            )
            db.add(lesson)
            added += 1
        except Exception as e:
            errors.append(f"Строка {row_idx}: {str(e)}")

    await db.commit()
    await log_action(db, "import_manual_lessons", {"filename": file.filename, "added": added, "errors": errors}, request)
    return {
        "message": "Загрузка завершена",
        "added": added,
        "errors": errors
    }