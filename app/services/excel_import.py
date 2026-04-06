import io
from typing import List, Dict, Any
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models.reference import (
    Group, Teacher, Audience, Building, Calendar, Faculty,
    UserType, Subdivision, Role, Permission, TimeSlot, Subject
)
from app.schemas.group import GroupCreate
from app.schemas.teacher import TeacherCreate
from app.schemas.audience import AudienceCreate
from app.schemas.building import BuildingCreate
from app.schemas.calendar import CalendarCreate
from app.schemas.faculty import FacultyCreate
from app.schemas.user_type import UserTypeCreate
from app.schemas.subdivision import SubdivisionCreate
from app.schemas.role import RoleCreate
from app.schemas.permission import PermissionCreate
from app.schemas.time_slot import TimeSlotCreate
from app.schemas.subject import SubjectCreate
from datetime import datetime, date

# ---------- Общая вспомогательная функция для чтения Excel ----------
def _read_excel(file_content: bytes, expected_columns: List[str]) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=io.BytesIO(file_content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_idx = {col: idx for idx, col in enumerate(headers) if col in expected_columns}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_data = {}
        for col_name, idx in col_idx.items():
            row_data[col_name] = row[idx]
        rows.append(row_data)
    return rows

# ---------- Группы ----------
async def parse_groups_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['name', 'faculty_id', 'student_count'])

async def validate_groups_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    faculty_ids = {row['faculty_id'] for row in rows if row.get('faculty_id')}
    existing_faculties = set()
    if faculty_ids:
        result = await db.execute(select(Faculty.id).where(Faculty.id.in_(faculty_ids)))
        existing_faculties = {r[0] for r in result.all()}
    for idx, row in enumerate(rows, start=2):
        try:
            validated = GroupCreate(
                name=row.get('name'),
                faculty_id=row.get('faculty_id'),
                student_count=row.get('student_count')
            )
            if validated.faculty_id and validated.faculty_id not in existing_faculties:
                errors.append({"row": idx, "error": f"Факультет id={validated.faculty_id} не найден"})
                continue
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_groups_from_validated(valid: List[GroupCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Group).where(Group.name == data.name)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.faculty_id = data.faculty_id
            existing.student_count = data.student_count
            updated += 1
        else:
            db.add(Group(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Преподаватели ----------
async def parse_teachers_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['login', 'name', 'url'])

async def validate_teachers_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            validated = TeacherCreate(
                login=row.get('login'),
                name=row.get('name'),
                url=row.get('url')
            )
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_teachers_from_validated(valid: List[TeacherCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Teacher).where(Teacher.login == data.login)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.name = data.name
            existing.url = data.url
            updated += 1
        else:
            db.add(Teacher(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Аудитории ----------
async def parse_audiences_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['name', 'building_id', 'capacity', 'type', 'is_active'])

async def validate_audiences_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    building_ids = {row['building_id'] for row in rows if row.get('building_id')}
    existing_buildings = set()
    if building_ids:
        result = await db.execute(select(Building.id).where(Building.id.in_(building_ids)))
        existing_buildings = {r[0] for r in result.all()}
    for idx, row in enumerate(rows, start=2):
        try:
            validated = AudienceCreate(
                name=row.get('name'),
                building_id=row.get('building_id'),
                capacity=row.get('capacity'),
                type=row.get('type'),
                is_active=row.get('is_active', True)
            )
            if validated.building_id and validated.building_id not in existing_buildings:
                errors.append({"row": idx, "error": f"Здание id={validated.building_id} не найдено"})
                continue
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_audiences_from_validated(valid: List[AudienceCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Audience).where(Audience.name == data.name)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.building_id = data.building_id
            existing.capacity = data.capacity
            existing.type = data.type
            existing.is_active = data.is_active
            updated += 1
        else:
            db.add(Audience(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Здания ----------
async def parse_buildings_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['name', 'address'])

async def validate_buildings_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            validated = BuildingCreate(
                name=row.get('name'),
                address=row.get('address')
            )
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_buildings_from_validated(valid: List[BuildingCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Building).where(Building.name == data.name)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.address = data.address
            updated += 1
        else:
            db.add(Building(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Календарь ----------
async def parse_calendar_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['date', 'is_working_day', 'week_type', 'description'])

async def validate_calendar_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            # Преобразование даты из Excel (может быть числом или строкой)
            raw_date = row.get('date')
            if isinstance(raw_date, datetime):
                parsed_date = raw_date.date()
            elif isinstance(raw_date, date):
                parsed_date = raw_date
            else:
                parsed_date = datetime.strptime(str(raw_date), "%Y-%m-%d").date()
            validated = CalendarCreate(
                date=parsed_date,
                is_working_day=row.get('is_working_day', True),
                week_type=row.get('week_type'),
                description=row.get('description')
            )
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_calendar_from_validated(valid: List[CalendarCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Calendar).where(Calendar.date == data.date)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.is_working_day = data.is_working_day
            existing.week_type = data.week_type
            existing.description = data.description
            updated += 1
        else:
            db.add(Calendar(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Факультеты ----------
async def parse_faculties_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['name', 'display_name', 'short_display_name'])

async def validate_faculties_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            validated = FacultyCreate(
                name=row.get('name'),
                display_name=row.get('display_name'),
                short_display_name=row.get('short_display_name')
            )
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_faculties_from_validated(valid: List[FacultyCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Faculty).where(Faculty.name == data.name)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.display_name = data.display_name
            existing.short_display_name = data.short_display_name
            updated += 1
        else:
            db.add(Faculty(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Типы пользователей ----------
async def parse_user_types_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['code', 'name', 'description'])

async def validate_user_types_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            validated = UserTypeCreate(
                code=row.get('code'),
                name=row.get('name'),
                description=row.get('description')
            )
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_user_types_from_validated(valid: List[UserTypeCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(UserType).where(UserType.code == data.code)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.name = data.name
            existing.description = data.description
            updated += 1
        else:
            db.add(UserType(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Подразделения ----------
async def parse_subdivisions_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['name'])

async def validate_subdivisions_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            validated = SubdivisionCreate(name=row.get('name'))
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_subdivisions_from_validated(valid: List[SubdivisionCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Subdivision).where(Subdivision.name == data.name)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            updated += 1
        else:
            db.add(Subdivision(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Роли ----------
async def parse_roles_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['role_name', 'description'])

async def validate_roles_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            validated = RoleCreate(
                role_name=row.get('role_name'),
                description=row.get('description')
            )
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_roles_from_validated(valid: List[RoleCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Role).where(Role.role_name == data.role_name)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.description = data.description
            updated += 1
        else:
            db.add(Role(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Права ----------
async def parse_permissions_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['permission_code', 'permission_name', 'description'])

async def validate_permissions_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            validated = PermissionCreate(
                permission_code=row.get('permission_code'),
                permission_name=row.get('permission_name'),
                description=row.get('description')
            )
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_permissions_from_validated(valid: List[PermissionCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Permission).where(Permission.permission_code == data.permission_code)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.permission_name = data.permission_name
            existing.description = data.description
            updated += 1
        else:
            db.add(Permission(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Временные слоты ----------
async def parse_time_slots_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['slot_number', 'name', 'start_time', 'end_time', 'duration_minutes', 'break_after_minutes', 'is_active'])

async def validate_time_slots_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            validated = TimeSlotCreate(
                slot_number=row.get('slot_number'),
                name=row.get('name'),
                start_time=row.get('start_time'),
                end_time=row.get('end_time'),
                duration_minutes=row.get('duration_minutes'),
                break_after_minutes=row.get('break_after_minutes'),
                is_active=row.get('is_active', True)
            )
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_time_slots_from_validated(valid: List[TimeSlotCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(TimeSlot).where(TimeSlot.slot_number == data.slot_number)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            existing.name = data.name
            existing.start_time = data.start_time
            existing.end_time = data.end_time
            existing.duration_minutes = data.duration_minutes
            existing.break_after_minutes = data.break_after_minutes
            existing.is_active = data.is_active
            updated += 1
        else:
            db.add(TimeSlot(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}

# ---------- Дисциплины (Subject) ----------
async def parse_subjects_excel(file_content: bytes) -> List[Dict[str, Any]]:
    return _read_excel(file_content, ['name'])

async def validate_subjects_data(rows: List[Dict[str, Any]], db: AsyncSession):
    valid = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            validated = SubjectCreate(name=row.get('name'))
            valid.append(validated)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    return valid, errors

async def save_subjects_from_validated(valid: List[SubjectCreate], db: AsyncSession):
    added = 0
    updated = 0
    for data in valid:
        stmt = select(Subject).where(Subject.name == data.name)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            updated += 1
        else:
            db.add(Subject(**data.model_dump()))
            added += 1
    await db.commit()
    return {"added": added, "updated": updated}